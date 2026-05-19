"""
InsureTrack — CYS Caloyeras Insurance Portfolio Manager
Rose Garden Edition · v11.0
"""
import json, os, io
from datetime import date, datetime
from collections import defaultdict

import streamlit as st
import plotly.graph_objects as go

try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False

try:
    from fpdf import FPDF
    HAS_FPDF = True
except ImportError:
    HAS_FPDF = False

# ─── Page config ─────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="InsureTrack · CYS",
    page_icon="🌹",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─── Rose Garden palette ─────────────────────────────────────────────────────
R = dict(
    bg        = "#FDF6F0",
    plum      = "#4A1942",
    plum_mid  = "#6B2C61",
    rose      = "#D4547A",
    rose_lt   = "#F2C4CE",
    rose_pale = "#FDE8EE",
    text_d    = "#3B0764",
    text_m    = "#9D8090",
    text_l    = "#C4A0B0",
    white     = "#FFFFFF",
    amber     = "#F59E0B",
    amber_lt  = "#FFFBEB",
    green     = "#22C55E",
    green_lt  = "#F0FDF4",
    red       = "#EF4444",
    red_lt    = "#FEE2E2",
    gray_lt   = "#F3F4F6",
)

# ─── Global CSS ──────────────────────────────────────────────────────────────
st.markdown(
    '<script src="https://cdn.jsdelivr.net/npm/@phosphor-icons/web@2.1.1/dist/index.js"></script>',
    unsafe_allow_html=True
)

st.markdown(f"""
<style>
html, body, [data-testid="stAppViewContainer"],
[data-testid="stMain"], .main {{
    background: {R['bg']} !important;
    color: {R['text_d']};
    font-family: 'Inter', 'Segoe UI', sans-serif;
}}
[data-testid="stSidebar"] {{ background: {R['plum']} !important; }}
[data-testid="stSidebar"] * {{ color: #EDD8E8 !important; }}
#MainMenu, footer, header {{ visibility: hidden; }}
[data-testid="stToolbar"] {{ display: none; }}
.block-container {{ padding: 1.5rem 2rem 2rem; max-width: 1400px; }}
::-webkit-scrollbar {{ width: 6px; height: 6px; }}
::-webkit-scrollbar-thumb {{ background: {R['rose_lt']}; border-radius: 3px; }}
.stButton > button {{
    background: {R['rose']} !important; color: {R['white']} !important;
    border: none !important; border-radius: 8px !important;
    padding: 0.45rem 1.2rem !important; font-weight: 600 !important;
    transition: all .18s;
}}
.stButton > button:hover {{
    background: {R['plum']} !important; transform: translateY(-1px);
    box-shadow: 0 4px 12px rgba(74,25,66,.25) !important;
}}
/* Sidebar nav — ghost style, not solid rose */
[data-testid="stSidebar"] .stButton > button {{
    background: transparent !important; color: #C4A0B0 !important;
    border: none !important; border-radius: 6px !important;
    text-align: left !important; font-weight: 500 !important;
    padding: 0.45rem 0.8rem !important;
    justify-content: flex-start !important;
    transform: none !important; box-shadow: none !important;
}}
[data-testid="stSidebar"] .stButton > button:hover {{
    background: rgba(255,255,255,.10) !important; color: #fff !important;
    transform: none !important; box-shadow: none !important;
}}
.stTextInput > div > div > input,
.stSelectbox > div > div {{
    border: 1.5px solid {R['rose_lt']} !important;
    border-radius: 8px !important;
    background: {R['white']} !important; color: {R['text_d']} !important;
}}
.stTextInput > div > div > input:focus {{
    border-color: {R['rose']} !important;
    box-shadow: 0 0 0 3px rgba(212,84,122,.15) !important;
}}
.stNumberInput > div > div > input,
.stNumberInput input,
div[data-testid="stNumberInput"] input,
div[data-testid="stNumberInput"] > div > input {{
    border: 1.5px solid {R['rose_lt']} !important;
    border-radius: 8px !important;
    background: {R['white']} !important; color: {R['text_d']} !important;
}}
.stTextArea > div > div > textarea,
div[data-testid="stTextArea"] textarea {{
    border: 1.5px solid {R['rose_lt']} !important;
    border-radius: 8px !important;
    background: {R['white']} !important; color: {R['text_d']} !important;
}}
/* Stepper buttons on number inputs */
.stNumberInput button {{
    background: {R['rose_pale']} !important; color: {R['rose']} !important;
    border: none !important; border-radius: 6px !important;
    transform: none !important; box-shadow: none !important;
    padding: 0.2rem 0.5rem !important; font-size: .9rem !important;
}}
.stTabs [data-baseweb="tab-list"] {{
    background: {R['white']}; border-radius: 10px; padding: 4px;
    border: 1.5px solid {R['rose_lt']}; gap: 4px;
}}
.stTabs [data-baseweb="tab"] {{
    border-radius: 8px !important; color: {R['text_m']} !important;
    font-weight: 600 !important; padding: 0.5rem 1.2rem !important;
}}
.stTabs [aria-selected="true"] {{
    background: {R['plum']} !important; color: {R['white']} !important;
}}
.kpi-tile {{
    background: {R['white']}; border: 1.5px solid {R['rose_lt']};
    border-radius: 12px; padding: 1.1rem 1.3rem; text-align: center;
}}
.kpi-tile .kpi-num {{
    font-size: 2rem; font-weight: 800; line-height: 1.1;
}}
.kpi-tile .kpi-lbl {{
    font-size: 0.68rem; font-weight: 700; letter-spacing: .07em;
    text-transform: uppercase; margin-top: 4px; color: {R['text_m']};
}}
.sec-hdr {{
    font-size: 0.72rem; font-weight: 700; letter-spacing: .1em;
    text-transform: uppercase; color: {R['text_m']};
    border-bottom: 2px solid {R['rose_lt']};
    padding-bottom: 6px; margin: 1.2rem 0 0.9rem;
    display: flex; align-items: center; gap: 6px;
}}
/* Richer expanders */
.streamlit-expanderHeader {{
    background: {R['white']} !important;
    border: 1.5px solid {R['rose_lt']} !important;
    border-radius: 10px !important;
    padding: .55rem 1rem !important;
    font-weight: 600 !important; color: {R['text_d']} !important;
    transition: background .15s;
}}
.streamlit-expanderHeader:hover {{
    background: {R['rose_pale']} !important;
    border-color: {R['rose']} !important;
}}
.streamlit-expanderContent {{
    border: 1.5px solid {R['rose_lt']} !important;
    border-top: none !important; border-radius: 0 0 10px 10px !important;
    background: {R['white']} !important; padding: .8rem 1rem !important;
}}
/* Better sidebar logo area */
[data-testid="stSidebar"] .stMarkdown h1,
[data-testid="stSidebar"] .stMarkdown p {{ color: #EDD8E8 !important; }}
/* Page header style */
.page-hdr {{
    font-size: 1.55rem; font-weight: 800; color: {R['text_d']};
    margin-bottom: .15rem; display: flex; align-items: center; gap: .5rem;
}}
.page-sub {{
    font-size: .83rem; color: {R['text_m']}; margin-bottom: 1.1rem;
}}
.badge {{
    display: inline-block; padding: 2px 10px;
    border-radius: 20px; font-size: 0.72rem; font-weight: 700;
}}
.badge-active  {{ background: {R['green_lt']};  color: #15803D; }}
.badge-quote   {{ background: {R['amber_lt']};  color: #92400E; }}
.badge-uninsu  {{ background: {R['red_lt']};    color: #991B1B; }}
.badge-verify  {{ background: {R['amber_lt']};  color: #92400E; }}
.badge-ext     {{ background: {R['gray_lt']};   color: {R['text_m']}; }}
.rg-card-sm {{
    background: {R['white']}; border: 1.5px solid {R['rose_lt']};
    border-radius: 10px; padding: 0.9rem 1.1rem;
}}
.badge-lapsed  {{ background: #F3E8FF; color: #6B21A8; }}
.ph, .ph-bold, .ph-fill {{ vertical-align: -2px; }}
</style>
""", unsafe_allow_html=True)


# ─── Phosphor Icons helper ────────────────────────────────────────
def phi(name: str, size: int = 16, color: str = "", weight: str = "regular") -> str:
    """Return a Phosphor icon <i> tag. weight: regular|bold|fill"""
    cls = f"ph-{weight} ph-{name}" if weight != "regular" else f"ph ph-{name}"
    style_str = f"font-size:{size}px;" + (f"color:{color};" if color else "")
    return f'<i class="{cls}" style="{style_str}" aria-hidden="true"></i>'


# ═══════════════════════════════════════════════════════════════════
#  DATA
# ═══════════════════════════════════════════════════════════════════
DATA_PATH = os.path.join(os.path.dirname(__file__), "data", "caloyeras", "portfolio.json")

@st.cache_data(ttl=300)
def load_data(path=DATA_PATH):
    try:
        with open(path) as f:
            raw = json.load(f)
        raw.setdefault("properties", [])
        raw.setdefault("policies", [])
        raw.setdefault("auto_policies", [])
        raw.setdefault("portfolio_name", "CYS Caloyeras")
        raw.setdefault("as_of_date", "")
        # Pre-compute unique policies once at load time
        seen, uniq = set(), []
        for p in raw["policies"]:
            pn = p.get("policy_number", "")
            if pn not in seen:
                seen.add(pn)
                uniq.append(p)
        raw["_policies_unique"] = uniq
        return raw
    except FileNotFoundError:
        return {"properties": [], "policies": [], "auto_policies": [],
                "_policies_unique": [], "portfolio_name": "CYS Caloyeras", "as_of_date": ""}
    except json.JSONDecodeError as e:
        st.error(f"portfolio.json is corrupted: {e}. Please restore from backup or re-import your data.")
        st.stop()

def save_data(data, path=DATA_PATH):
    tmp = path + ".tmp"
    with open(tmp, "w") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    os.replace(tmp, path)
    load_data.clear()

def _today():
    return date.today()

def _days_to(date_str):
    if not date_str:
        return None
    try:
        return (datetime.strptime(date_str, "%Y-%m-%d").date() - _today()).days
    except Exception:
        return None

def unique_policies(policies):
    seen, out = set(), []
    for p in policies:
        pn = p.get("policy_number", "")
        if pn not in seen:
            seen.add(pn)
            out.append(p)
    return out

def total_premium(policies):
    return sum(p.get("premium") or 0 for p in unique_policies(policies))

# ─── Lapsed policy detector ───────────────────────────────────────
def compute_coverage_status(prop, prop_policies):
    """Returns effective display status — detects lapsed policies automatically."""
    manual = prop.get("coverage_status", "Uninsured")
    if manual != "Active":
        return manual
    if not prop_policies:
        return "Lapsed"
    try:
        active_pols = [
            p for p in prop_policies
            if p.get("expiration_date") and
            datetime.strptime(p["expiration_date"], "%Y-%m-%d").date() >= _today()
        ]
    except Exception:
        return manual
    return "Active" if active_pols else "Lapsed"

# ─── Data completeness score ──────────────────────────────────────
_COMPLETENESS_FIELDS = ["address", "city", "state", "owner", "units", "type"]

def data_completeness(prop, prop_policies):
    filled = sum(1 for f in _COMPLETENESS_FIELDS if prop.get(f))
    has_policy = bool(prop_policies)
    has_expiry = any(p.get("expiration_date") for p in prop_policies)
    total = len(_COMPLETENESS_FIELDS) + 2
    return round((filled + int(has_policy) + int(has_expiry)) / total * 100)

def coverage_badge(status):
    if status == "Active":
        return '<span class="badge badge-active"><i class="ph ph-check-circle" style="font-size:11px;margin-right:3px;"></i>Active</span>'
    if status == "Lapsed":
        return '<span class="badge badge-lapsed"><i class="ph ph-warning" style="font-size:11px;margin-right:3px;"></i>Lapsed</span>'
    if status == "Quote":
        return '<span class="badge badge-quote"><i class="ph ph-clock" style="font-size:11px;margin-right:3px;"></i>Quote</span>'
    if status == "Uninsured":
        return '<span class="badge badge-uninsu"><i class="ph ph-x-circle" style="font-size:11px;margin-right:3px;"></i>Uninsured</span>'
    if "Verify" in status and "External" not in status:
        return '<span class="badge badge-verify"><i class="ph ph-question" style="font-size:11px;margin-right:3px;"></i>Verify</span>'
    if "External" in status:
        return '<span class="badge badge-ext"><i class="ph ph-user-switch" style="font-size:11px;margin-right:3px;"></i>Ext Owner</span>'
    return f'<span class="badge badge-ext">{status}</span>'


# ─── City → lat/lon lookup (no API key needed) ────────────────────
CITY_COORDS = {
    # ── California ───────────────────────────────────────────────
    "Los Angeles":       (34.0522, -118.2437),
    "Santa Monica":      (34.0195, -118.4912),
    "Inglewood":         (33.9617, -118.3531),
    "Mission Viejo":     (33.6000, -117.6720),
    "Culver City":       (34.0211, -118.3965),
    "Torrance":          (33.8358, -118.3406),
    "Long Beach":        (33.7701, -118.1937),
    "Pasadena":          (34.1478, -118.1445),
    "Burbank":           (34.1808, -118.3090),
    "Glendale":          (34.1425, -118.2551),
    "Compton":           (33.8958, -118.2201),
    "Hawthorne":         (33.9164, -118.3526),
    "Gardena":           (33.8883, -118.3089),
    "Carson":            (33.8317, -118.2820),
    "Lawndale":          (33.8872, -118.3526),
    "El Segundo":        (33.9192, -118.4165),
    "Manhattan Beach":   (33.8847, -118.4109),
    "Redondo Beach":     (33.8492, -118.3884),
    "Hermosa Beach":     (33.8622, -118.3995),
    "Beverly Hills":     (34.0736, -118.4004),
    "West Hollywood":    (34.0900, -118.3617),
    "Encino":            (34.1520, -118.5011),
    "Reseda":            (34.1992, -118.5353),
    "Sherman Oaks":      (34.1503, -118.4514),
    "Van Nuys":          (34.1858, -118.4487),
    "North Hollywood":   (34.1872, -118.3790),
    "Canoga Park":       (34.1997, -118.5984),
    "Woodland Hills":    (34.1683, -118.6059),
    "Chatsworth":        (34.2572, -118.6043),
    "Northridge":        (34.2305, -118.5353),
    "Granada Hills":     (34.2739, -118.5028),
    "West Hills":        (34.2009, -118.6420),
    "Calabasas":         (34.1347, -118.6601),
    "Thousand Oaks":     (34.1706, -118.8376),
    "Simi Valley":       (34.2694, -118.7815),
    "Ventura":           (34.2748, -119.2290),
    "Oxnard":            (34.1975, -119.1771),
    "Camarillo":         (34.2164, -119.0376),
    "Moorpark":          (34.2858, -118.8820),
    "Agoura Hills":      (34.1431, -118.7614),
    "Pomona":            (34.0553, -117.7500),
    "Ontario":           (34.0633, -117.6509),
    "Rancho Cucamonga":  (34.1064, -117.5931),
    "San Bernardino":    (34.1083, -117.2898),
    "Riverside":         (33.9806, -117.3755),
    "Corona":            (33.8753, -117.5664),
    "Anaheim":           (33.8366, -117.9143),
    "Santa Ana":         (33.7455, -117.8677),
    "Irvine":            (33.6846, -117.8265),
    "Huntington Beach":  (33.6595, -118.0000),
    "Garden Grove":      (33.7743, -117.9378),
    "Orange":            (33.7879, -117.8531),
    "Fullerton":         (33.8704, -117.9242),
    "Buena Park":        (33.8675, -117.9981),
    "Westminster":       (33.7592, -117.9939),
    "Costa Mesa":        (33.6411, -117.9187),
    "Newport Beach":     (33.6189, -117.9289),
    "Laguna Hills":      (33.5878, -117.7156),
    "Laguna Niguel":     (33.5225, -117.7076),
    "San Clemente":      (33.4269, -117.6120),
    "San Juan Capistrano": (33.5017, -117.6625),
    "Dana Point":        (33.4669, -117.6981),
    "Aliso Viejo":       (33.5769, -117.7261),
    "Lake Forest":       (33.6469, -117.6892),
    "Foothill Ranch":    (33.6800, -117.6650),
    "Tustin":            (33.7458, -117.8261),
    "Yorba Linda":       (33.8886, -117.8131),
    "Brea":              (33.9167, -117.9003),
    "La Habra":          (33.9319, -117.9462),
    "Cerritos":          (33.8583, -118.0647),
    "Lakewood":          (33.8536, -118.1339),
    "Bellflower":        (33.8817, -118.1170),
    "Norwalk":           (33.9022, -118.0814),
    "Downey":            (33.9400, -118.1328),
    "Paramount":         (33.8894, -118.1597),
    "South Gate":        (33.9547, -118.2120),
    "Lynwood":           (33.9303, -118.2114),
    "Watts":             (33.9267, -118.2461),
    "Montebello":        (34.0161, -118.1136),
    "Commerce":          (33.9961, -118.1595),
    "Bell":              (33.9775, -118.1872),
    "Bell Gardens":      (33.9656, -118.1517),
    "Maywood":           (33.9872, -118.1878),
    "Huntington Park":   (33.9814, -118.2248),
    "Cudahy":            (33.9622, -118.1856),
    "Florence":          (33.9681, -118.2428),
    "Pico Rivera":       (33.9828, -118.0967),
    "Whittier":          (33.9792, -118.0328),
    "La Mirada":         (33.9017, -118.0122),
    "Santa Fe Springs":  (33.9428, -118.0753),
    "Artesia":           (33.8653, -118.0803),
    "Hawaiian Gardens":  (33.8322, -118.0714),
    "Azusa":             (34.1336, -117.9073),
    "Monrovia":          (34.1442, -117.9995),
    "Arcadia":           (34.1397, -118.0353),
    "Temple City":       (34.1067, -118.0567),
    "Rosemead":          (34.0806, -118.0728),
    "El Monte":          (34.0686, -118.0275),
    "West Covina":       (34.0686, -117.9394),
    "Covina":            (34.0900, -117.8903),
    "Baldwin Park":      (34.0853, -117.9608),
    "Industry":          (34.0153, -117.9628),
    "Walnut":            (34.0220, -117.8656),
    "Diamond Bar":       (34.0289, -117.8103),
    "Rowland Heights":   (33.9764, -117.8981),
    "La Puente":         (34.0219, -117.9500),
    "Hacienda Heights":  (33.9931, -117.9706),
    "Pico Union":        (34.0519, -118.2811),
    "Echo Park":         (34.0782, -118.2609),
    "Silver Lake":       (34.0869, -118.2717),
    "Highland Park":     (34.1076, -118.1897),
    "Eagle Rock":        (34.1395, -118.2087),
    "Glassell Park":     (34.1100, -118.2300),
    "Atwater Village":   (34.1147, -118.2653),
    "Los Feliz":         (34.1078, -118.2972),
    "Koreatown":         (34.0586, -118.3019),
    "Mid-City":          (34.0400, -118.3400),
    "Crenshaw":          (34.0100, -118.3328),
    "Leimert Park":      (34.0050, -118.3333),
    "Hyde Park":         (33.9800, -118.3278),
    "View Park":         (34.0000, -118.3500),
    "Windsor Hills":     (33.9900, -118.3400),
    "West Adams":        (34.0222, -118.3161),
    "Jefferson Park":    (34.0181, -118.3222),
    "Exposition Park":   (34.0167, -118.2878),
    "University Park":   (34.0200, -118.2850),
    "Vermont Square":    (34.0022, -118.2939),
    "Gramercy Park":     (34.0311, -118.3008),
    "Harvard Park":      (33.9850, -118.3058),
    "Florence-Firestone": (33.9681, -118.2428),
    "Willowbrook":       (33.9258, -118.2550),
    # ── Washington State ─────────────────────────────────────────
    "Seattle":        (47.6062, -122.3321),
    "Tacoma":         (47.2529, -122.4443),
    "Spokane":        (47.6588, -117.4260),
    "Bellevue":       (47.6101, -122.2015),
    "Renton":         (47.4799, -122.2171),
    "Kent":           (47.3809, -122.2348),
    "Auburn":         (47.3073, -122.2285),
    "Federal Way":    (47.3223, -122.3126),
    "Kirkland":       (47.6815, -122.2087),
    "Redmond":        (47.6740, -122.1215),
    "Everett":        (47.9790, -122.2021),
    "Marysville":     (48.0517, -122.1771),
    "Lynnwood":       (47.8209, -122.3151),
    "Edmonds":        (47.8107, -122.3776),
    "Shoreline":      (47.7543, -122.3416),
    "Burien":         (47.4704, -122.3468),
    "Tukwila":        (47.4740, -122.2612),
    "SeaTac":         (47.4435, -122.2987),
    "Des Moines":     (47.4015, -122.3243),
    "Puyallup":       (47.1854, -122.2929),
    "Lakewood WA":    (47.1718, -122.5185),
    "Olympia":        (47.0379, -122.9007),
    "Bremerton":      (47.5673, -122.6329),
    "Yakima":         (46.6021, -120.5059),
    "Kennewick":      (46.2112, -119.1372),
    "Richland":       (46.2857, -119.2845),
    "Pasco":          (46.2396, -119.1006),
    "Vancouver":      (45.6387, -122.6615),
    "Bellingham":     (48.7519, -122.4787),
    "Mount Vernon":   (48.4215, -122.3343),
    "Wenatchee":      (47.4235, -120.3103),
    "Ellensburg":     (46.9965, -120.5478),
    "Walla Walla":    (46.0646, -118.3430),
    "Pullman":        (46.7298, -117.1817),
    "Longview":       (46.1382, -122.9382),
    "Centralia":      (46.7165, -122.9543),
    "Port Angeles":   (48.1181, -123.4307),
    "Port Townsend":  (48.1170, -122.7601),
    "Anacortes":      (48.5126, -122.6127),
    "Oak Harbor":     (48.2937, -122.6429),
    "Poulsbo":        (47.7354, -122.6468),
    "Gig Harbor":     (47.3298, -122.5793),
    "Covington":      (47.3584, -122.1079),
    "Maple Valley":   (47.3665, -122.0454),
    "Black Diamond":  (47.3088, -122.0004),
    "Enumclaw":       (47.2018, -121.9918),
    "Bonney Lake":    (47.1779, -122.1801),
    "Sumner":         (47.2040, -122.2290),
    "Pacific":        (47.2640, -122.2526),
    "Milton":         (47.2529, -122.3148),
    "Fife":           (47.2376, -122.3590),
    "Edgewood":       (47.2285, -122.2945),
    "South Hill":     (47.1462, -122.2943),
    "Spanaway":       (47.1015, -122.4285),
    "Graham":         (47.0523, -122.2987),
    "University Place": (47.2126, -122.5435),
    "Steilacoom":     (47.1718, -122.5985),
    "DuPont":         (47.0979, -122.6271),
    "Orting":         (47.0993, -122.2024),
    "Eatonville":     (46.8712, -122.2641),
    "Carbonado":      (47.0843, -122.0546),
    "Wilkeson":       (47.0982, -122.0398),
    "Buckley":        (47.1596, -122.0257),
    "Greenwater":     (47.1337, -121.6338),
    "Ashford":        (46.7599, -122.0227),
    "Yelm":           (46.9415, -122.6143),
    "Roy":            (47.0120, -122.3776),
    "Anderson Island":(47.1543, -122.7048),
    "Vashon":         (47.4418, -122.4709),
    "Bainbridge Island": (47.6426, -122.5381),
    "Fox Island":     (47.2565, -122.6298),
    "Mercer Island":  (47.5707, -122.2221),
    "Medina":         (47.6243, -122.2265),
    "Clyde Hill":     (47.6293, -122.2110),
    "Yarrow Point":   (47.6482, -122.2123),
    "Hunts Point":    (47.6468, -122.2176),
    "Beaux Arts Village": (47.5943, -122.1990),
    "Newcastle":      (47.5293, -122.1601),
    "Sammamish":      (47.6163, -122.0355),
    "Issaquah":       (47.5301, -122.0326),
    "North Bend":     (47.4957, -121.7868),
    "Snoqualmie":     (47.5265, -121.8257),
    "Fall City":      (47.5615, -121.9035),
    "Carnation":      (47.6479, -121.9152),
    "Duvall":         (47.7426, -121.9843),
    "Monroe":         (47.8554, -121.9713),
    "Snohomish":      (47.9129, -122.0985),
    "Gold Bar":       (47.8521, -121.6924),
    "Index":          (47.8193, -121.5526),
    "Sultan":         (47.8654, -121.8124),
    "Startup":        (47.8526, -121.7468),
    "Skykomish":      (47.7101, -121.3563),
    "Leavenworth":    (47.5963, -120.6615),
    "Cashmere":       (47.5218, -120.4682),
    "Chelan":         (47.8393, -120.0182),
    "Manson":         (47.8837, -120.1568),
    "Entiat":         (47.6693, -120.2115),
    "Malaga":         (47.3751, -120.3032),
    "Quincy":         (47.2354, -119.8529),
    "Ephrata":        (47.3179, -119.5527),
    "Moses Lake":     (47.1301, -119.2779),
    "Othello":        (46.8268, -119.1712),
    "Ritzville":      (47.1265, -118.3791),
    "Davenport":      (47.6573, -118.1591),
    "Wilbur":         (47.7554, -118.7032),
    "Grand Coulee":   (47.9393, -118.9757),
    "Electric City":  (47.9260, -119.0135),
    "Coulee City":    (47.6107, -119.2910),
    "Soap Lake":      (47.3882, -119.4985),
    "Cheney":         (47.4873, -117.5782),
    "Deer Park":      (47.9554, -117.4757),
    "Mead":           (47.7690, -117.3535),
    "Medical Lake":   (47.5693, -117.6813),
    "Airway Heights": (47.6457, -117.5921),
    "Spokane Valley": (47.6732, -117.2394),
    "Liberty Lake":   (47.6693, -117.1040),
    "Millwood":       (47.6726, -117.2915),
    "Veradale":       (47.6565, -117.2035),
    "Greenacres":     (47.6640, -117.1610),
    "Otis Orchards":  (47.6943, -117.1029),
    "Newman Lake":    (47.7254, -117.0713),
    "Colbert":        (47.8582, -117.3682),
    "Chattaroy":      (47.8982, -117.3440),
    "Elk":            (47.9415, -117.2835),
    "Clayton":        (47.9921, -117.5593),
    "Ford":           (47.8726, -117.8007),
    "Springdale":     (48.0576, -117.7640),
    "Chewelah":       (48.2776, -117.7157),
    "Colville":       (48.5465, -117.9018),
    "Kettle Falls":   (48.6054, -118.0574),
    "Republic":       (48.6487, -118.7326),
    "Tonasket":       (48.7051, -119.4382),
    "Okanogan":       (48.3598, -119.5729),
    "Omak":           (48.4115, -119.5279),
    "Pateros":        (48.0526, -119.9018),
    "Bridgeport":     (47.9771, -119.6643),
    "Brewster":       (48.0951, -119.7779),
    "Twisp":          (48.3607, -120.1196),
    "Winthrop":       (48.4751, -120.1793),
    "Mazama":         (48.5951, -120.4024),
    "Concrete":       (48.5376, -121.7532),
    "Sedro-Woolley":  (48.5054, -122.2368),
    "Burlington":     (48.4751, -122.3296),
    "Bow":            (48.5426, -122.3793),
    "La Conner":      (48.3943, -122.4982),
    "Coupeville":     (48.2204, -122.6854),
    "Langley":        (48.0418, -122.4076),
    "Clinton":        (47.9715, -122.3593),
    "Freeland":       (48.0001, -122.5168),
    "Greenbank":      (48.1043, -122.5785),
    "Mukilteo":       (47.9490, -122.3040),
    "Mountlake Terrace": (47.7887, -122.3090),
    "Kenmore":        (47.7579, -122.2440),
    "Bothell":        (47.7623, -122.2054),
    "Mill Creek":     (47.8601, -122.2040),
    "Woodinville":    (47.7540, -122.1637),
    "Granite Falls":  (48.0840, -121.9724),
    "Arlington":      (48.1968, -122.1196),
    "Stanwood":       (48.2423, -122.3732),
    "Camano Island":  (48.1979, -122.4754),
    "Whidbey Island": (48.2204, -122.6854),
}

CHANGELOG_PATH = os.path.join(os.path.dirname(__file__), "data", "caloyeras", "changelog.json")

@st.cache_data(ttl=60)
def load_changelog():
    if os.path.exists(CHANGELOG_PATH):
        try:
            with open(CHANGELOG_PATH) as f:
                return json.load(f)
        except Exception:
            return []
    return []

def log_change(action, detail, user=None):
    changelog = load_changelog()
    entry = {
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "user": user or st.session_state.get("username", "system"),
        "action": action,
        "detail": detail,
    }
    changelog.insert(0, entry)
    changelog = changelog[:500]
    os.makedirs(os.path.dirname(CHANGELOG_PATH), exist_ok=True)
    tmp = CHANGELOG_PATH + ".tmp"
    with open(tmp, "w") as f:
        json.dump(changelog, f, indent=2, ensure_ascii=False)
    os.replace(tmp, CHANGELOG_PATH)
    load_changelog.clear()

def _get_coords(city, prop_id=""):
    base = CITY_COORDS.get(city)
    if not base:
        # Try partial match
        for k, v in CITY_COORDS.items():
            if city and k.lower().startswith(city.lower()[:4]):
                base = v
                break
    if not base:
        base = (34.0522, -118.2437)  # Default: Los Angeles
    # Deterministic jitter by prop_id so props don't stack
    seed = sum(ord(c) for c in str(prop_id)) if prop_id else 0
    lat_off = ((seed * 17) % 100 - 50) * 0.0015
    lon_off = ((seed * 31) % 100 - 50) * 0.0015
    return base[0] + lat_off, base[1] + lon_off


# ═══════════════════════════════════════════════════════════════════
#  LOGIN
# ═══════════════════════════════════════════════════════════════════
def _get_users():
    try:
        users_secret = st.secrets.get("users", {})
        if users_secret:
            return dict(users_secret)
    except Exception:
        pass
    return {"caloyeras": "cys2026", "admin": "insuretrack"}  # local dev fallback

USERS = _get_users()

def show_login():
    col = st.columns([1, 1.1, 1])[1]
    with col:
        st.markdown("<br><br>", unsafe_allow_html=True)
        st.markdown(f"""
        <div style="background:{R['white']};border:1.5px solid {R['rose_lt']};
             border-radius:18px;padding:2.5rem;text-align:center;
             box-shadow:0 8px 32px rgba(74,25,66,.12);">
          <div style="font-size:2rem;margin-bottom:.4rem;color:#D4547A;"><i class="ph ph-shield-star"></i></div>
          <div style="font-size:1.5rem;font-weight:800;color:{R['text_d']};">InsureTrack</div>
          <div style="font-size:.85rem;color:{R['text_m']};margin-bottom:1.8rem;">
            CYS Caloyeras · Portfolio Manager</div>
        </div>
        """, unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)
        user = st.text_input("user", placeholder="Username", label_visibility="collapsed",
                             key="li_user")
        pw   = st.text_input("pw", type="password", placeholder="Password",
                             label_visibility="collapsed", key="li_pw")
        if st.button("Sign In →", use_container_width=True):
            if USERS.get(user) == pw:
                st.session_state.logged_in  = True
                st.session_state.username   = user
                st.session_state.page       = "dashboard"
                st.session_state.login_time = __import__("time").time()
                st.rerun()
            else:
                st.error("Invalid credentials.")
        st.markdown(
            f"<div style='text-align:center;font-size:.72rem;color:{R['text_l']};"
            f"margin-top:.6rem;'>v11.0 · Rose Garden</div>",
            unsafe_allow_html=True
        )


# ═══════════════════════════════════════════════════════════════════
#  SIDEBAR
# ═══════════════════════════════════════════════════════════════════
NAV = [
    ("squares-four",    "Dashboard",   "dashboard"),
    ("magnifying-glass","Search",      "search"),
    ("map-pin",         "Map",         "map"),
    ("check-square",    "Tasks",       "tasks"),
    ("clock-countdown", "Renewals",    "queue"),
    ("file-text",       "Policies",    "policies"),
    ("buildings",       "Properties",  "properties"),
    ("car",             "Auto",        "auto"),
    ("calendar",        "Calendar",    "calendar"),
    ("chart-line",      "Analytics",   "analytics"),
    ("plus-circle",     "Add / Edit",  "add"),
    ("upload-simple",   "Import",      "import_data"),
    ("newspaper",       "Reports",     "reports"),
    ("gear",            "Settings",    "settings"),
]

def show_sidebar(data):
    with st.sidebar:
        st.markdown(f"""
        <div style="padding:1rem .4rem .8rem;">
          <div style="font-size:1.2rem;font-weight:800;color:{R['white']};display:flex;align-items:center;gap:8px;">
            <i class="ph ph-shield-star" style="font-size:1.3rem;"></i> InsureTrack</div>
          <div style="font-size:.7rem;color:{R['text_l']};margin-top:2px;">
            {data.get('portfolio_name','CYS Caloyeras')}</div>
        </div>
        <hr style="border:none;border-top:1px solid rgba(255,255,255,.12);margin:.2rem 0 .6rem;">
        """, unsafe_allow_html=True)

        cur = st.session_state.get("page", "dashboard")
        for icon_name, label, key in NAV:
            btn_label = f"{phi(icon_name, 14)}  {label}"
            if st.button(btn_label, key=f"nav_{key}", use_container_width=True):
                st.session_state.page = key
                st.rerun()

        props    = data["properties"]
        active_n = sum(1 for p in props if p.get("coverage_status") == "Active")
        uninsu_n = sum(1 for p in props if p.get("coverage_status") == "Uninsured")
        verify_n = sum(1 for p in props if "Verify" in p.get("coverage_status",""))

        st.markdown(f"""
        <hr style="border:none;border-top:1px solid rgba(255,255,255,.12);margin:.8rem 0 .5rem;">
        <div style="padding:.2rem .4rem;font-size:.7rem;color:{R['text_l']};">
          <i class="ph ph-package" style="font-size:12px;"></i> {len(props)} props &middot; {sum(p.get('units') or 0 for p in props)} units<br>
          <i class="ph ph-check-circle" style="font-size:12px;"></i> {active_n} insured &nbsp; <i class="ph ph-warning" style="font-size:12px;"></i> {uninsu_n + verify_n} gaps
        </div>
        <hr style="border:none;border-top:1px solid rgba(255,255,255,.12);margin:.5rem 0 .4rem;">
        <div style="padding:.1rem .4rem;font-size:.66rem;color:{R['text_l']};">
          v11.0 · Rose Garden · {data.get('as_of_date','')}
        </div>
        """, unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════════
#  DASHBOARD
# ═══════════════════════════════════════════════════════════════════
def page_dashboard(data):
    props    = data["properties"]
    policies = data["policies"]

    active_n  = sum(1 for p in props if p.get("coverage_status") == "Active")
    uninsu_n  = sum(1 for p in props if p.get("coverage_status") == "Uninsured")
    verify_n  = sum(1 for p in props if "Verify" in p.get("coverage_status",""))
    prem_tot  = total_premium(policies)
    tot_units = sum(p.get("units") or 0 for p in props)

    st.markdown(f"""
    <div class="page-hdr">{phi("squares-four",24)} Dashboard</div>
    <div class="page-sub">{data.get('portfolio_name','')} &nbsp;·&nbsp; Updated: {data.get('as_of_date','')}</div>
    """, unsafe_allow_html=True)

    # Build policy-per-prop lookup for lapsed detection
    pol_by_prop = {}
    for pol in policies:
        pol_by_prop.setdefault(pol.get("prop_id",""), []).append(pol)

    # KPIs — white cards with left-border accent + colored numbers
    quote_n   = sum(1 for p in props if p.get("coverage_status") == "Quote")
    lapsed_n  = sum(1 for p in props
                    if compute_coverage_status(p, pol_by_prop.get(p["prop_id"], [])) == "Lapsed")
    avg_comp  = round(sum(data_completeness(p, pol_by_prop.get(p["prop_id"], [])) for p in props) / len(props)) if props else 0
    kpi_data = [
        (str(len(props)),          "Total Properties", R['text_d'],  R['plum']),
        (str(active_n),            "Insured Active",   "#16a34a",    "#16a34a"),
        (str(uninsu_n + lapsed_n), "Uninsured / Lapsed","#dc2626",  "#dc2626"),
        (str(verify_n + quote_n),  "Verify / Quote",   "#d97706",    "#d97706"),
        (f"${prem_tot:,.0f}",      "Annual Premium",   R['rose'],    R['rose']),
        (f"{avg_comp}%",           "Data Quality",     R['plum'],    R['plum_mid']),
    ]
    cols = st.columns(6)
    for col, (num, lbl, color, accent) in zip(cols, kpi_data):
        col.markdown(f"""
        <div style="background:{R['white']};border:1.5px solid {R['rose_lt']};
             border-left:4px solid {accent};border-radius:12px;
             padding:1rem 1.1rem;text-align:center;
             box-shadow:0 2px 8px rgba(74,25,66,.06);">
          <div style="font-size:1.85rem;font-weight:800;line-height:1.1;color:{color};">{num}</div>
          <div style="font-size:0.65rem;font-weight:700;letter-spacing:.08em;
               text-transform:uppercase;margin-top:5px;color:{R['text_m']};">{lbl}</div>
        </div>""", unsafe_allow_html=True)

    # ── Priority Alert Banner ──────────────────────────────────────
    urgent_pols = [p for p in unique_policies(policies)
                   if (_days_to(p.get("expiration_date")) or 999) <= 60
                   and (_days_to(p.get("expiration_date")) or 999) >= 0]
    gap_count   = uninsu_n + verify_n + quote_n
    alerts = []
    if urgent_pols:
        _pol_word = 'policy' if len(urgent_pols) == 1 else 'policies'
        alerts.append(
            f'<i class="ph ph-clock" style="font-size:14px;color:#b45309;vertical-align:-2px;"></i>'
            f' <b>{len(urgent_pols)} {_pol_word}</b> expiring in less than 60 days'
        )
    if uninsu_n:
        alerts.append(
            f'<i class="ph ph-x-circle" style="font-size:14px;color:#dc2626;vertical-align:-2px;"></i>'
            f' <b>{uninsu_n} properties</b> uninsured — action required'
        )
    if verify_n:
        alerts.append(f'<i class="ph ph-warning" style="font-size:14px;color:#b45309;vertical-align:-2px;"></i> <b>{verify_n} properties</b> pending verification')

    st.markdown("<div style='height:.6rem'></div>", unsafe_allow_html=True)
    if alerts:
        alert_html = "&nbsp;&nbsp;·&nbsp;&nbsp;".join(alerts)
        st.markdown(f"""
        <div style="background:linear-gradient(90deg,#fff1f2,#fffbeb);
             border:1.5px solid #fca5a5;border-radius:12px;
             padding:.8rem 1.4rem;margin-bottom:.8rem;
             display:flex;align-items:center;gap:12px;flex-wrap:wrap;">
          <i class="ph ph-siren" style="font-size:1.2rem;color:#dc2626;vertical-align:-3px;flex-shrink:0;"></i>
          <span style="font-size:.82rem;color:#7f1d1d;flex:1;">{alert_html}</span>
        </div>
        """, unsafe_allow_html=True)
        # CTA buttons
        cta1, cta2, cta3 = st.columns([1, 1, 3])
        with cta1:
            if st.button("View uninsured properties", key="cta_uninsu", use_container_width=True):
                st.session_state.page = "properties"
                st.session_state["prop_open_gaps"] = True
                st.rerun()
        with cta2:
            if st.button("View expiring policies", key="cta_pol", use_container_width=True):
                st.session_state.page = "policies"
                st.rerun()
    else:
        st.markdown(f"""
        <div style="background:{R['green_lt']};border:1.5px solid #86efac;
             border-radius:12px;padding:.7rem 1.4rem;margin-bottom:.8rem;">
          <span style="font-size:.85rem;color:#14532d;">
            <i class="ph ph-check-circle" style="font-size:15px;color:#16a34a;vertical-align:-2px;"></i> <b>All clear</b> — no urgent alerts in the portfolio</span>
        </div>
        """, unsafe_allow_html=True)

    st.markdown("<div style='height:.4rem'></div>", unsafe_allow_html=True)

    left, right = st.columns([1.05, 1], gap="large")

    # ── Renewals ──
    with left:
        st.markdown(f'<div class="sec-hdr"><i class="ph ph-calendar" style="font-size:13px;margin-right:4px;"></i>Upcoming Renewals</div>', unsafe_allow_html=True)
        pol_props_map = defaultdict(list)
        for p in policies:
            if p.get("prop_id"):
                pol_props_map[p["policy_number"]].append(p["prop_id"])

        renewals = sorted(
            [p for p in unique_policies(policies) if p.get("expiration_date")],
            key=lambda x: x["expiration_date"]
        )
        for pol in renewals[:12]:
            days = _days_to(pol["expiration_date"])
            if days is None:
                continue
            color   = R['red'] if days <= 60 else (R['amber'] if days <= 180 else R['green'])
            bg      = R['red_lt'] if days <= 60 else (R['amber_lt'] if days <= 180 else R['green_lt'])
            pids    = " · ".join(pol_props_map.get(pol["policy_number"], ["—"]))
            carrier = (pol.get("carrier") or "").split("/")[0].strip()[:24]
            urgency = "◉" if days <= 60 else ("○" if days <= 180 else "●")
            st.markdown(f"""
            <div style="display:flex;align-items:center;gap:10px;padding:.55rem .9rem;
                 background:{bg};border-radius:8px;
                 border-left:3px solid {color};margin-bottom:5px;">
              <div style="min-width:48px;text-align:center;">
                <div style="font-weight:800;font-size:.95rem;color:{color};">{days}d</div>
                <div style="font-size:.75rem;">{urgency}</div>
              </div>
              <div style="flex:1;overflow:hidden;">
                <div style="font-size:.8rem;font-weight:700;color:{R['text_d']};
                     white-space:nowrap;overflow:hidden;text-overflow:ellipsis;">
                  {pol['policy_number']}</div>
                <div style="font-size:.7rem;color:{R['text_m']};">{carrier} · {pids}</div>
              </div>
              <div style="font-size:.7rem;color:{R['text_m']};white-space:nowrap;text-align:right;">
                {pol['expiration_date']}<br>
                <span style="color:{color};font-weight:600;">${pol.get('premium',0):,.0f}</span>
              </div>
            </div>""", unsafe_allow_html=True)

        # Coverage Gaps quick list
        st.markdown(f'<div class="sec-hdr"><i class="ph ph-x-circle" style="font-size:14px;vertical-align:-2px;margin-right:5px;"></i>Coverage Gaps — Priority</div>',
                    unsafe_allow_html=True)
        gaps = sorted(
            [p for p in props if p.get("coverage_status") == "Uninsured"],
            key=lambda p: -(p.get("units") or 0)
        )
        for prop in gaps[:7]:
            units = prop.get("units") or 0
            pri   = "HIGH" if units >= 10 else ("MED" if units >= 4 else "LOW")
            pri_col = R['red'] if units >= 10 else (R['amber'] if units >= 4 else "#ca8a04")
            owner = (prop.get("owner") or "—")[:22]
            st.markdown(f"""
            <div style="display:flex;align-items:center;gap:10px;padding:.55rem .9rem;
                 background:{R['red_lt']};border-radius:8px;
                 border-left:3px solid {pri_col};margin-bottom:5px;">
              <div style="font-size:.7rem;font-weight:800;color:{pri_col};min-width:52px;">{pri}</div>
              <div style="flex:1;overflow:hidden;">
                <div style="font-size:.8rem;font-weight:700;color:{R['text_d']};
                     white-space:nowrap;overflow:hidden;text-overflow:ellipsis;">
                  {prop['prop_id']} · {prop.get('nickname') or prop['address']}</div>
                <div style="font-size:.7rem;color:{R['text_m']};">
                  {prop['city']} · {units} unids · {owner}</div>
              </div>
              <div style="font-size:.7rem;color:{R['red']};font-weight:700;
                   white-space:nowrap;">Uninsured</div>
            </div>""", unsafe_allow_html=True)
        if len(gaps) > 7:
            st.markdown(f"""<div style="font-size:.75rem;color:{R['text_m']};
                padding:.3rem .8rem;">
                + {len(gaps)-7} more → go to Properties / Coverage Gaps</div>""",
                unsafe_allow_html=True)

    # ── Charts ──
    with right:
        st.markdown(f'<div class="sec-hdr"><i class="ph ph-buildings" style="font-size:14px;vertical-align:-2px;margin-right:5px;"></i>Carrier Concentration</div>',
                    unsafe_allow_html=True)
        carrier_prem = defaultdict(float)
        for pol in unique_policies(policies):
            c = (pol.get("carrier") or "Unknown").split("/")[0].strip()
            c = c.replace("Insurance Exchange","").replace("Insurance Co","").replace("Ins.","").strip()
            if c.endswith(" "): c = c.strip()
            carrier_prem[c] += pol.get("premium") or 0

        carriers = sorted(carrier_prem.items(), key=lambda x: -x[1])
        labels   = [c[0][:20] for c in carriers]
        values   = [c[1] for c in carriers]
        palette  = [R['plum'], R['rose'], R['amber'], "#7C3AED", "#0EA5E9",
                    "#10B981", "#F97316", R['plum_mid'], R['text_m'], "#64748B"]

        fig = go.Figure(go.Pie(
            labels=labels, values=values, hole=.54,
            marker_colors=palette[:len(labels)],
            textfont_size=10,
            hovertemplate="<b>%{label}</b><br>$%{value:,.0f}<extra></extra>",
        ))
        fig.update_layout(
            height=250, margin=dict(l=0, r=0, t=0, b=0),
            paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
            showlegend=True,
            legend=dict(font_size=9, x=1.02, y=.5, bgcolor="rgba(0,0,0,0)"),
            annotations=[dict(
                text=f"<b>${prem_tot/1000:.0f}K</b>",
                x=.5, y=.5, font_size=14, showarrow=False,
                font_color=R['text_d']
            )]
        )
        st.plotly_chart(fig, use_container_width=True, config={"displayModeBar": False})

        # Renewal bar chart
        st.markdown(f'<div class="sec-hdr"><i class="ph ph-calendar" style="font-size:14px;vertical-align:-2px;margin-right:5px;"></i>Renewals — Next 12 Months</div>',
                    unsafe_allow_html=True)
        today = _today()
        buckets = defaultdict(float)
        for pol in unique_policies(policies):
            exp = pol.get("expiration_date")
            if not exp: continue
            d     = datetime.strptime(exp, "%Y-%m-%d").date()
            delta = (d.year - today.year) * 12 + d.month - today.month
            if 0 <= delta < 12:
                buckets[d.strftime("%b '%y")] += pol.get("premium") or 0

        if buckets:
            mb = sorted(buckets.items(), key=lambda x: datetime.strptime(x[0], "%b '%y"))
            months = [m[0] for m in mb]
            prems  = [m[1] for m in mb]
            fig2 = go.Figure(go.Bar(
                x=months, y=prems,
                marker_color=[R['red'] if p > 30000 else R['rose'] for p in prems],
                hovertemplate="%{x}<br>$%{y:,.0f}<extra></extra>",
            ))
            fig2.update_layout(
                height=190, margin=dict(l=0, r=0, t=5, b=0),
                paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                xaxis=dict(showgrid=False, tickfont_size=9),
                yaxis=dict(showgrid=True, gridcolor="#F0E8EC",
                           tickformat="$,.0f", tickfont_size=8),
            )
            st.plotly_chart(fig2, use_container_width=True,
                            config={"displayModeBar": False})

        # Coverage status pie
        st.markdown(f'<div class="sec-hdr"><i class="ph ph-shield-check" style="font-size:13px;margin-right:4px;"></i>Coverage Distribution</div>',
                    unsafe_allow_html=True)
        status_counts = defaultdict(int)
        for p in props:
            s = p.get("coverage_status","")
            if "External" in s or "Verify" in s:
                status_counts["Verify / External"] += 1
            elif s:
                status_counts[s] += 1
            else:
                status_counts["Unknown"] += 1

        fig3 = go.Figure(go.Pie(
            labels=list(status_counts.keys()),
            values=list(status_counts.values()),
            hole=.4,
            marker_colors=[R['green'], "#991B1B", R['amber'], R['plum'], R['text_m']][:len(status_counts)],
            textfont_size=10,
            hovertemplate="<b>%{label}</b><br>%{value} props<extra></extra>",
        ))
        fig3.update_layout(
            height=190, margin=dict(l=0, r=0, t=0, b=0),
            paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
            showlegend=True,
            legend=dict(font_size=9, x=1.02, y=.5, bgcolor="rgba(0,0,0,0)"),
        )
        st.plotly_chart(fig3, use_container_width=True, config={"displayModeBar": False})


# ═══════════════════════════════════════════════════════════════════
#  POLICIES
# ═══════════════════════════════════════════════════════════════════
def page_policies(data):
    props    = data["properties"]
    policies = data["policies"]

    prem_tot = total_premium(policies)
    st.markdown(f"""
    <div style="font-size:1.5rem;font-weight:800;color:{R['text_d']};margin-bottom:.2rem;">
      <i class="ph ph-file-text" style="font-size:1.4rem;vertical-align:-4px;margin-right:6px;"></i>Policies</div>
    <div style="font-size:.85rem;color:{R['text_m']};margin-bottom:1.2rem;">
      {len(unique_policies(policies))} unique policies · ${prem_tot:,.2f} annual premium</div>
    """, unsafe_allow_html=True)

    pol_props_map = defaultdict(list)
    for p in policies:
        if p.get("prop_id"):
            pol_props_map[p["policy_number"]].append(p["prop_id"])

    fcol1, fcol2, fcol3 = st.columns([2, 1.2, 1.2])
    with fcol1:
        search = st.text_input("s", placeholder="Search policy # or carrier…",
                               label_visibility="collapsed", key="pol_q")
    with fcol2:
        sf = st.selectbox("st", ["All Statuses","Active","Quote","Expired"],
                          label_visibility="collapsed", key="pol_sf")
    with fcol3:
        so = st.selectbox("so", ["Sort: Expiration","Sort: Premium ↓","Sort: Carrier"],
                          label_visibility="collapsed", key="pol_so")

    uniq = unique_policies(policies)
    if search:
        q = search.lower()
        uniq = [p for p in uniq
                if q in p["policy_number"].lower()
                or q in (p.get("carrier") or "").lower()
                or q in (p.get("agency") or "").lower()]
    if sf != "All Statuses":
        uniq = [p for p in uniq if p.get("status") == sf]
    if "Premium" in so:
        uniq.sort(key=lambda x: -(x.get("premium") or 0))
    elif "Carrier" in so:
        uniq.sort(key=lambda x: x.get("carrier") or "")
    else:
        uniq.sort(key=lambda x: x.get("expiration_date") or "")

    st.caption(f"{len(uniq)} policies shown")

    for pol in uniq:
        days    = _days_to(pol.get("expiration_date"))
        status  = pol.get("status","")
        s_color = (R['green'] if status == "Active" else
                   R['amber'] if status == "Quote" else R['red'])
        s_icon  = ("●" if status == "Active" else
                   "◑" if status == "Quote" else
                   "○" if status in ("Expired","Cancelled") else "◌")
        days_str = f"{days}d" if days is not None else "—"
        days_col = (R['red'] if (days or 999) <= 60 else
                    R['amber'] if (days or 999) <= 180 else R['green'])
        days_icon = ('<i class="ph ph-clock" style="font-size:13px;color:#dc2626;vertical-align:-2px;"></i>' if (days or 999) <= 60 else
                     '<i class="ph ph-lightning" style="font-size:13px;color:#d97706;vertical-align:-2px;"></i>' if (days or 999) <= 180 else "")
        pids     = " · ".join(pol_props_map.get(pol["policy_number"], ["—"]))
        carrier  = (pol.get("carrier") or "")
        prem     = pol.get("premium") or 0

        with st.expander(
            f"{s_icon}  {pol['policy_number']}  ·  {carrier[:28]}  ·  ${prem:,.0f}"
            f"  {days_icon} {days_str}",
            expanded=False
        ):
            r1, r2, r3, r4 = st.columns(4)
            r1.markdown(
                f"**Status**<br>"
                f"<span style='color:{s_color};font-weight:700;'>{status}</span>",
                unsafe_allow_html=True)
            r2.markdown(
                f"**Expires**<br>"
                f"<span style='color:{days_col};font-weight:700;'>"
                f"{pol.get('expiration_date','—')} ({days_str})</span>",
                unsafe_allow_html=True)
            r3.markdown(f"**Premium**<br><b>${prem:,.2f}</b>", unsafe_allow_html=True)
            bldg = pol.get("building_limit")
            r4.markdown(f"**Bldg Limit**<br><b>${bldg:,.0f}</b>" if bldg else
                        "**Bldg Limit**<br>—", unsafe_allow_html=True)

            st.divider()
            d1, d2, d3 = st.columns(3)
            d1.markdown(f"**Type**<br>{pol.get('policy_type','—')}", unsafe_allow_html=True)
            d2.markdown(f"**Agency**<br>{pol.get('agency','—')}", unsafe_allow_html=True)
            d3.markdown(f"**Properties**<br>{pids}", unsafe_allow_html=True)

            d4, d5, d6 = st.columns(3)
            d4.markdown(f"**Ded AOP**<br>{pol.get('ded_aop','—')}", unsafe_allow_html=True)
            d5.markdown(f"**Ded Water**<br>{pol.get('ded_water','—')}", unsafe_allow_html=True)
            d6.markdown(f"**Habitability**<br>{pol.get('habitability','—')}", unsafe_allow_html=True)

            if pol.get("business_income"):
                st.caption(f"BI: {pol['business_income']}")
            if pol.get("liability"):
                st.caption(f"Liability: {pol['liability']}")
            if pol.get("notes"):
                st.warning(pol['notes'][:120])


# ═══════════════════════════════════════════════════════════════════
#  PROPERTIES
# ═══════════════════════════════════════════════════════════════════
def page_properties(data):
    props    = data["properties"]
    policies = data["policies"]

    pol_lookup = {}
    pol_all_lookup: dict = {}          # prop_id → [all policies]
    for p in policies:
        pid = p.get("prop_id")
        if pid:
            if pid not in pol_lookup:
                pol_lookup[pid] = p
            pol_all_lookup.setdefault(pid, []).append(p)

    tot_units = sum(p.get("units") or 0 for p in props)
    gap_count = sum(1 for p in props
                    if p.get("coverage_status") not in ("Active",""))

    st.markdown(f"""
    <div style="font-size:1.5rem;font-weight:800;color:{R['text_d']};margin-bottom:.2rem;">
      <i class="ph ph-buildings" style="font-size:1.4rem;vertical-align:-4px;margin-right:6px;"></i>Properties</div>
    <div style="font-size:.85rem;color:{R['text_m']};margin-bottom:1.2rem;">
      {len(props)} properties · {tot_units} units</div>
    """, unsafe_allow_html=True)

    # Auto-open gaps tab if navigated from dashboard alert
    default_tab = 1 if st.session_state.pop("prop_open_gaps", False) else 0

    tab_all, tab_gaps = st.tabs([
        f"  All ({len(props)})  ",
        f"  Coverage Gaps ({gap_count})  ",
    ])

    def prop_card(prop, pol, all_pols=None):
        """Render a single property card with drill-down timeline and quick-edit."""
        status   = prop.get("coverage_status","")
        days     = _days_to(pol.get("expiration_date")) if pol else None
        days_col = (R['red'] if (days or 999) <= 60 else
                    R['amber'] if (days or 999) <= 180 else R['green'])
        badge    = coverage_badge(status)
        pid      = prop["prop_id"]

        # Status icon for expander title (HTML not supported in title)
        status_icon = ("●" if status == "Active" else
                       "✗" if status == "Uninsured" else
                       "◎"  if status == "Quote" else
                       "▲" if "Verify" in status or "External" in status else "○")

        with st.expander(
            f"{status_icon} {pid} · {prop.get('nickname') or prop['address']} · {prop['city']}"
            f" ({prop.get('units') or 0}u)",
            expanded=False
        ):
            # ── Quick-Edit toggle ──────────────────────────────────────────────
            qe_key = f"qe_{pid}"
            qe_active = st.session_state.get(qe_key, False)
            btn_cols = st.columns([8, 1])
            with btn_cols[1]:
                if st.button("Edit" if not qe_active else "Close",
                             key=f"qe_btn_{pid}", use_container_width=True):
                    st.session_state[qe_key] = not qe_active
                    st.rerun()

            if qe_active:
                # ── Inline Quick-Edit form ─────────────────────────────────────
                st.markdown(
                    f"<div style='background:{R['bg']};border:1px solid {R['rose_lt']};"
                    f"border-radius:8px;padding:.8rem 1rem;margin-bottom:.6rem;'>"
                    f"<b>Quick Edit — {pid}</b></div>",
                    unsafe_allow_html=True
                )
                qe1, qe2 = st.columns(2)
                new_status = qe1.selectbox(
                    "Coverage Status", ["Active","Uninsured","Quote","Verify / External"],
                    index=["Active","Uninsured","Quote","Verify / External"].index(
                        status if status in ["Active","Uninsured","Quote","Verify / External"] else "Active"),
                    key=f"qe_status_{pid}"
                )
                new_action = qe2.selectbox(
                    "Action Status",
                    ["None","Needs Attention","In Progress","Pending Quote","Done"],
                    index=["None","Needs Attention","In Progress","Pending Quote","Done"].index(
                        prop.get("action_status","None") if prop.get("action_status","None")
                        in ["None","Needs Attention","In Progress","Pending Quote","Done"] else "None"),
                    key=f"qe_action_{pid}"
                )
                new_notes = st.text_area("Notes", value=prop.get("notes",""),
                                         key=f"qe_notes_{pid}", height=60)
                sv1, sv2 = st.columns(2)
                if sv1.button("Save", key=f"qe_save_{pid}", type="primary",
                              use_container_width=True):
                    for p in data["properties"]:
                        if p["prop_id"] == pid:
                            p["coverage_status"] = new_status
                            p["action_status"]   = new_action
                            p["notes"]           = new_notes
                            break
                    save_data(data)
                    _log_change("edit", f"Quick-edited {pid}")
                    st.session_state[qe_key] = False
                    st.success(f"{pid} updated.")
                    st.rerun()
                if sv2.button("Cancel", key=f"qe_cancel_{pid}", use_container_width=True):
                    st.session_state[qe_key] = False
                    st.rerun()
                return   # don't render rest of card while editing

            # ── Standard card view ────────────────────────────────────────────
            r1, r2, r3 = st.columns(3)
            r1.markdown(f"**Address**<br>{prop['address']}, {prop['city']} {prop.get('zip','')}",
                        unsafe_allow_html=True)
            r2.markdown(f"**Owner**<br>{prop.get('owner','—')}", unsafe_allow_html=True)
            r3.markdown(f"**Coverage**<br>{badge}", unsafe_allow_html=True)

            r4, r5, r6 = st.columns(3)
            r4.markdown(f"**Units:** {prop.get('units') or '—'}")
            sqft = prop.get('sqft')
            r5.markdown(f"**Sq Ft:** {f'{sqft:,}' if sqft else '—'}")
            r6.markdown(f"**Type:** {prop.get('type','—')}")

            if pol:
                st.divider()
                p1, p2, p3 = st.columns(3)
                p1.markdown(f"**Policy #**<br>{pol.get('policy_number','—')}",
                            unsafe_allow_html=True)
                p2.markdown(f"**Carrier**<br>{(pol.get('carrier') or '—')[:30]}",
                            unsafe_allow_html=True)
                p3.markdown(
                    f"**Expires**<br>"
                    f"<span style='color:{days_col};font-weight:700;'>"
                    f"{pol.get('expiration_date','—')}"
                    f"{f' ({days}d)' if days is not None else ''}</span>",
                    unsafe_allow_html=True
                )
                p4, p5 = st.columns(2)
                prem = pol.get('premium') or 0
                bldg = pol.get('building_limit') or 0
                p4.markdown(f"**Premium:** ${prem:,.2f}")
                p5.markdown(f"**Bldg Limit:** ${bldg:,.0f}")

            # ── Policy Timeline drill-down ─────────────────────────────────────
            if all_pols and len(all_pols) > 0:
                timeline_pols = sorted(
                    all_pols,
                    key=lambda p: p.get("effective_date") or p.get("expiration_date") or "",
                    reverse=True
                )
                with st.expander(
                    f"Policy history ({len(timeline_pols)} {'policy' if len(timeline_pols)==1 else 'policies'})",
                    expanded=False
                ):
                    tl_rows = []
                    for tp in timeline_pols:
                        d_exp  = tp.get("expiration_date","")
                        d_eff  = tp.get("effective_date","")
                        d_left = _days_to(d_exp)
                        status_dot = "●" if (d_left is not None and d_left >= 0) else "○"
                        dot_color  = R['green'] if (d_left is not None and d_left >= 0) else R['text_m']
                        tl_rows.append(
                            f"<tr style='border-bottom:1px solid {R['rose_lt']};'>"
                            f"<td style='padding:4px 6px;font-size:.73rem;color:{dot_color};'>{status_dot}</td>"
                            f"<td style='padding:4px 6px;font-size:.73rem;font-weight:600;'>{tp.get('policy_number','—')}</td>"
                            f"<td style='padding:4px 6px;font-size:.73rem;color:{R['text_m']};'>{(tp.get('carrier') or '—')[:22]}</td>"
                            f"<td style='padding:4px 6px;font-size:.73rem;'>{d_eff or '—'}</td>"
                            f"<td style='padding:4px 6px;font-size:.73rem;color:{R['red'] if (d_left or 999)<60 else R['text_d']};'>{d_exp or '—'}</td>"
                            f"<td style='padding:4px 6px;font-size:.73rem;'>${(tp.get('premium') or 0):,.0f}</td>"
                            f"</tr>"
                        )
                    st.markdown(
                        f"<table style='width:100%;border-collapse:collapse;'>"
                        f"<thead><tr style='background:{R['rose_lt']};'>"
                        f"<th style='padding:4px 6px;font-size:.7rem;text-align:left;'></th>"
                        f"<th style='padding:4px 6px;font-size:.7rem;text-align:left;'>Policy #</th>"
                        f"<th style='padding:4px 6px;font-size:.7rem;text-align:left;'>Carrier</th>"
                        f"<th style='padding:4px 6px;font-size:.7rem;text-align:left;'>Effective</th>"
                        f"<th style='padding:4px 6px;font-size:.7rem;text-align:left;'>Expires</th>"
                        f"<th style='padding:4px 6px;font-size:.7rem;text-align:left;'>Premium</th>"
                        f"</tr></thead><tbody>{''.join(tl_rows)}</tbody></table>",
                        unsafe_allow_html=True
                    )

            action = prop.get("action_status","")
            if action and action != "None":
                action_color = ("#dc2626" if action == "Needs Attention" else
                                "#d97706" if action == "In Progress" else
                                "#d97706" if action == "Pending Quote" else
                                "#16a34a")
                st.markdown(f"""
                <div style="display:inline-block;padding:3px 12px;border-radius:20px;
                     background:{action_color}20;border:1px solid {action_color};
                     color:{action_color};font-size:.72rem;font-weight:700;margin:.4rem 0;">
                  <i class="ph ph-bell" style="font-size:12px;vertical-align:-1px;"></i> {action}
                </div>""", unsafe_allow_html=True)
            if prop.get("notes"):
                st.info(prop['notes'])
            if prop.get("mortgagee"):
                st.caption(f"Mortgagee: {prop['mortgagee']}")

    # ── All tab ──
    with tab_all:
        c1, c2, c3 = st.columns([2, 1.2, 1.4])
        with c1:
            srch = st.text_input("s2", placeholder="Search by address, city, owner, ID…",
                                 label_visibility="collapsed", key="pr_srch")
        with c2:
            sf2 = st.selectbox("sf2", ["All","Active","Uninsured","Quote","Verify"],
                               label_visibility="collapsed", key="pr_sf")
        with c3:
            so2 = st.selectbox("so2",
                ["Sort: Default","Units ↓","City A→Z","Owner A→Z"],
                label_visibility="collapsed", key="pr_so")

        filt = list(props)
        if srch:
            q = srch.lower()
            filt = [p for p in filt
                    if q in p.get("address","").lower()
                    or q in p.get("city","").lower()
                    or q in (p.get("owner") or "").lower()
                    or q in p.get("prop_id","").lower()
                    or q in (p.get("nickname") or "").lower()]
        if sf2 != "All":
            if sf2 == "Verify":
                filt = [p for p in filt if "Verify" in p.get("coverage_status","")]
            else:
                filt = [p for p in filt if p.get("coverage_status") == sf2]

        if "Units" in so2:
            filt.sort(key=lambda p: -(p.get("units") or 0))
        elif "City" in so2:
            filt.sort(key=lambda p: p.get("city",""))
        elif "Owner" in so2:
            filt.sort(key=lambda p: (p.get("owner") or ""))

        # Quick stats summary bar
        vis_active  = sum(1 for p in filt if p.get("coverage_status") == "Active")
        vis_uninsu  = sum(1 for p in filt if p.get("coverage_status") == "Uninsured")
        vis_other   = len(filt) - vis_active - vis_uninsu
        vis_units   = sum(p.get("units") or 0 for p in filt)
        st.markdown(f"""
        <div style="display:flex;gap:16px;padding:.5rem .8rem;background:{R['white']};
             border:1px solid {R['rose_lt']};border-radius:8px;margin-bottom:.6rem;
             font-size:.75rem;flex-wrap:wrap;">
          <span><b>{len(filt)}</b> properties</span>
          <span style="color:#16a34a;"><i class="ph ph-check-circle" style="font-size:13px;vertical-align:-2px;"></i> <b>{vis_active}</b> insured</span>
          <span style="color:#dc2626;"><i class="ph ph-x-circle" style="font-size:13px;vertical-align:-2px;"></i> <b>{vis_uninsu}</b> uninsured</span>
          {("<span style='color:#d97706;'><i class='ph ph-warning' style='font-size:13px;vertical-align:-2px;'></i> <b>" + str(vis_other) + "</b> other</span>") if vis_other else ""}
          <span style="margin-left:auto;color:{R['text_m']};"><i class="ph ph-buildings" style="font-size:13px;vertical-align:-2px;"></i> <b>{vis_units}</b> units</span>
        </div>
        """, unsafe_allow_html=True)

        # ── Bulk Status Editor ────────────────────────────────────────────
        bulk_key = "bulk_sel_props"
        if bulk_key not in st.session_state:
            st.session_state[bulk_key] = set()

        # Bulk action bar (shown when items are selected)
        selected_ids = st.session_state[bulk_key]
        if selected_ids:
            bulk_bar_cols = st.columns([3, 2, 1, 1])
            with bulk_bar_cols[0]:
                st.markdown(
                    f"<div style='padding:.4rem .8rem;background:{R['plum']}18;"
                    f"border:1px solid {R['plum']}40;border-radius:8px;"
                    f"font-size:.82rem;font-weight:700;color:{R['plum']};'>"
                    f"<i class='ph ph-check-square' style='vertical-align:-2px;margin-right:5px;'></i>"
                    f"{len(selected_ids)} selected</div>",
                    unsafe_allow_html=True
                )
            with bulk_bar_cols[1]:
                bulk_new_status = st.selectbox(
                    "bulk_status", ["Active","Uninsured","Quote","Verify / External"],
                    label_visibility="collapsed", key="bulk_status_sel"
                )
            with bulk_bar_cols[2]:
                if st.button("Apply", key="bulk_apply", type="primary",
                             use_container_width=True):
                    for p in data["properties"]:
                        if p["prop_id"] in selected_ids:
                            p["coverage_status"] = bulk_new_status
                    save_data(data)
                    _log_change("edit", f"Bulk status → {bulk_new_status} for {len(selected_ids)} props")
                    st.session_state[bulk_key] = set()
                    st.success(f"Updated {len(selected_ids)} properties to {bulk_new_status}.")
                    st.rerun()
            with bulk_bar_cols[3]:
                if st.button("Clear", key="bulk_clear", use_container_width=True):
                    st.session_state[bulk_key] = set()
                    st.rerun()

        # Render property cards with select checkbox
        for prop in filt:
            pid = prop["prop_id"]
            cb_col, card_col = st.columns([0.04, 0.96])
            with cb_col:
                checked = st.checkbox("", value=(pid in selected_ids),
                                      key=f"bulk_cb_{pid}", label_visibility="collapsed")
                if checked:
                    st.session_state[bulk_key].add(pid)
                else:
                    st.session_state[bulk_key].discard(pid)
            with card_col:
                prop_card(prop, pol_lookup.get(pid),
                          all_pols=pol_all_lookup.get(pid, []))

    # ── Coverage Gaps tab ──
    with tab_gaps:
        gap_props = [p for p in props if p.get("coverage_status") not in ("Active","")]
        uninsu  = sorted([p for p in gap_props if p.get("coverage_status") == "Uninsured"],
                         key=lambda p: -(p.get("units") or 0))
        verify  = [p for p in gap_props if "Verify" in p.get("coverage_status","")]
        ext     = [p for p in gap_props if "External" in p.get("coverage_status","")]
        qt      = [p for p in gap_props if p.get("coverage_status") == "Quote"]

        g1, g2, g3, g4 = st.columns(4)
        for col, n, lbl, num_color in [
            (g1, len(uninsu), "Uninsured", "#dc2626"),
            (g2, sum(p.get("units") or 0 for p in uninsu), "Uninsured Units", "#dc2626"),
            (g3, len(verify) + len(ext), "Need Verify", "#d97706"),
            (g4, len(qt), "Unbound Quote", R['plum']),
        ]:
            col.markdown(f"""
            <div class="kpi-tile">
              <div class="kpi-num" style="color:{num_color};">{n}</div>
              <div class="kpi-lbl">{lbl}</div>
            </div>""", unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)

        if uninsu:
            st.markdown(f'<div class="sec-hdr"><i class="ph ph-x-circle" style="font-size:14px;vertical-align:-2px;margin-right:5px;"></i>Uninsured — No Policy On File</div>',
                        unsafe_allow_html=True)
            for prop in uninsu:
                units = prop.get("units") or 0
                pri   = "HIGH" if units >= 10 else ("MED" if units >= 4 else "LOW")
                st.markdown(f"""
                <div style="display:flex;align-items:center;gap:12px;padding:.6rem .9rem;
                     background:{R['red_lt']};border-radius:8px;
                     border-left:3px solid {R['red']};margin-bottom:5px;">
                  <div style="font-size:.78rem;font-weight:800;min-width:60px;">{pri}</div>
                  <div style="flex:1;">
                    <div style="font-size:.85rem;font-weight:700;color:{R['text_d']};">
                      {prop['prop_id']} · {prop['address']}</div>
                    <div style="font-size:.72rem;color:{R['text_m']};">
                      {prop['city']} · {units} units · {prop.get('owner','—')}</div>
                  </div>
                  <div style="font-size:.72rem;color:#991B1B;font-weight:700;">No policy</div>
                </div>""", unsafe_allow_html=True)

        if verify or ext:
            st.markdown(f'<div class="sec-hdr"><i class="ph ph-warning" style="font-size:13px;margin-right:4px;"></i>Verify / External Owner</div>',
                        unsafe_allow_html=True)
            for prop in verify + ext:
                status = prop.get("coverage_status","")
                st.markdown(f"""
                <div style="display:flex;align-items:center;gap:12px;padding:.55rem .9rem;
                     background:{R['amber_lt']};border-radius:8px;
                     border-left:3px solid {R['amber']};margin-bottom:5px;">
                  <div style="flex:1;">
                    <div style="font-size:.85rem;font-weight:700;color:{R['text_d']};">
                      {prop['prop_id']} · {prop['address']}</div>
                    <div style="font-size:.72rem;color:{R['text_m']};">
                      {prop['city']} · {prop.get('units') or 0} units</div>
                  </div>
                  <div style="font-size:.72rem;color:#92400E;font-weight:600;
                       max-width:200px;text-align:right;">{status}</div>
                </div>""", unsafe_allow_html=True)
                if prop.get("notes"):
                    st.caption(prop['notes'])

        if qt:
            st.markdown(f'<div class="sec-hdr">◎ Quote — Not Yet Bound</div>',
                        unsafe_allow_html=True)
            for prop in qt:
                pol = pol_lookup.get(prop["prop_id"], {})
                st.markdown(f"""
                <div style="padding:.55rem .9rem;background:{R['amber_lt']};
                     border-radius:8px;border-left:3px solid {R['amber']};margin-bottom:5px;">
                  <div style="font-size:.85rem;font-weight:700;color:{R['text_d']};">
                    {prop['prop_id']} · {prop['address']}</div>
                  <div style="font-size:.72rem;color:{R['text_m']};">
                    {prop['city']} · {pol.get('policy_number','—')} ·
                    {pol.get('carrier','—')}</div>
                </div>""", unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════════
#  AUTO
# ═══════════════════════════════════════════════════════════════════
def page_auto(data):
    autos = data.get("auto_policies", [])
    total_ap = sum(a.get("premium") or 0 for a in autos)

    st.markdown(f"""
    <div style="font-size:1.5rem;font-weight:800;color:{R['text_d']};margin-bottom:.2rem;">
      <i class="ph ph-car" style="font-size:1.4rem;vertical-align:-4px;margin-right:6px;"></i>Auto Policies</div>
    <div style="font-size:.85rem;color:{R['text_m']};margin-bottom:1rem;">
      {len(autos)} policies · ${total_ap:,.2f} total premium</div>
    """, unsafe_allow_html=True)

    # Summary KPIs
    urgent_auto = sum(1 for a in autos if (_days_to(a.get("expiration_date")) or 999) <= 60)
    a1, a2, a3, a4 = st.columns(4)
    for col, num, lbl, color in [
        (a1, len(autos),           "Total Policies", R['text_d']),
        (a2, f"${total_ap:,.0f}",  "Total Premium",  R['rose']),
        (a3, urgent_auto or "—",   "Expiring < 60d", "#dc2626" if urgent_auto else R['text_m']),
        (a4, len(set(a.get("state","") for a in autos if a.get("state"))),
             "States",              R['plum']),
    ]:
        col.markdown(f"""
        <div class="kpi-tile">
          <div class="kpi-num" style="color:{color};font-size:1.5rem;">{num}</div>
          <div class="kpi-lbl">{lbl}</div>
        </div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    for auto in autos:
        days     = _days_to(auto.get("expiration_date"))
        days_str = f"{days}d" if days is not None else "—"
        days_col = (R['red'] if (days or 999) <= 60 else
                    R['amber'] if (days or 999) <= 180 else R['green'])
        days_icon = "(!)" if (days or 999) <= 60 else ("~" if (days or 999) <= 180 else "")
        prem    = auto.get("premium") or 0
        vehicle = (auto.get("vehicles") or auto.get("vehicle") or "")
        veh_short = vehicle[:25] if vehicle else auto.get("state","")

        with st.expander(
            f"{auto.get('insured','')[:24]}  ·  {veh_short}  ·  "
            f"${prem:,.0f}  ·  {days_icon} {days_str}",
            expanded=False
        ):
            r1, r2, r3 = st.columns(3)
            r1.markdown(f"**Insured**<br>{auto.get('insured','—')}", unsafe_allow_html=True)
            r2.markdown(f"**Carrier**<br>{auto.get('carrier','—')}", unsafe_allow_html=True)
            r3.markdown(f"**Agency**<br>{auto.get('agency','—')}", unsafe_allow_html=True)

            r4, r5, r6 = st.columns(3)
            r4.markdown(f"**Effective:** {auto.get('effective_date','—')}")
            r5.markdown(
                f"**Expires:** <span style='color:{days_col};font-weight:700;'>"
                f"{auto.get('expiration_date','—')} ({days_str})</span>",
                unsafe_allow_html=True)
            r6.markdown(f"**Premium:** ${prem:,.2f}")

            st.divider()
            v1, v2 = st.columns([1.5, 1])
            v1.markdown(f"**Vehicles**<br>{auto.get('vehicles','—')}", unsafe_allow_html=True)
            v2.markdown(f"**BI/PD**<br>{auto.get('bipd','—')}", unsafe_allow_html=True)

            v3, v4, v5 = st.columns(3)
            v3.markdown(f"**Comp Ded:** {auto.get('comp_ded','—')}")
            v4.markdown(f"**Coll Ded:** {auto.get('coll_ded','—')}")
            v5.markdown(f"**UM/UIM:** {auto.get('um_uim','—')}")

            if auto.get("vins"):
                st.caption(f"VINs: {auto['vins']}")
            if auto.get("pip_medpay"):
                st.caption(f"PIP/MedPay: {auto['pip_medpay']}")
            if auto.get("notes"):
                st.info(auto['notes'])


# ═══════════════════════════════════════════════════════════════════
#  REPORTS
# ═══════════════════════════════════════════════════════════════════
def page_reports(data):
    props    = data["properties"]
    policies = data["policies"]
    autos    = data.get("auto_policies", [])

    active_n = sum(1 for p in props if p.get("coverage_status") == "Active")
    uninsu_n = sum(1 for p in props if p.get("coverage_status") == "Uninsured")
    verify_n = sum(1 for p in props if "Verify" in p.get("coverage_status",""))
    prem_tot = total_premium(policies)
    auto_prem= sum(a.get("premium") or 0 for a in autos)
    tot_units= sum(p.get("units") or 0 for p in props)

    st.markdown(f"""
    <div style="font-size:1.5rem;font-weight:800;color:{R['text_d']};margin-bottom:.2rem;">
      <i class="ph ph-newspaper" style="font-size:1.4rem;vertical-align:-4px;margin-right:6px;"></i>Reports</div>
    <div style="font-size:.85rem;color:{R['text_m']};margin-bottom:1.2rem;">
      Executive summaries and data exports</div>
    """, unsafe_allow_html=True)

    today_str = _today().strftime("%B %d, %Y")
    report = f"""INSURETRACK — PORTFOLIO EXECUTIVE SUMMARY
{data.get('portfolio_name','CYS Caloyeras')}
Generated: {today_str}
{'='*60}

PORTFOLIO OVERVIEW
  Total Properties:     {len(props)}
  Total Units:          {tot_units}
  Annual Premium:       ${prem_tot:,.2f}  (property)
  Auto Premium:         ${auto_prem:,.2f}
  TOTAL PREMIUM:        ${prem_tot + auto_prem:,.2f}

COVERAGE STATUS
  Active (Insured):     {active_n}
  Quote (Unbound):      {sum(1 for p in props if p.get('coverage_status')=='Quote')}
  UNINSURED:            {uninsu_n}  ← ACTION REQUIRED
  Need Verification:    {verify_n}

UNINSURED PROPERTIES (sorted by unit count)
"""
    for i, prop in enumerate(
        sorted([p for p in props if p.get("coverage_status") == "Uninsured"],
               key=lambda p: -(p.get("units") or 0)), 1
    ):
        report += (f"  {i:2}. {prop['prop_id']} · {prop['address']}, "
                   f"{prop['city']} ({prop.get('units') or 0} units)\n")

    report += "\nRENEWALS NEXT 90 DAYS\n"
    near = sorted(
        [(p, d) for p in unique_policies(policies)
         if (d := _days_to(p.get("expiration_date"))) is not None and 0 <= d <= 90],
        key=lambda x: x[1]
    )
    if near:
        for pol, d in near:
            report += (f"  {d:3}d · {pol['policy_number']} · "
                       f"{(pol.get('carrier') or '')[:28]} · ${pol.get('premium') or 0:,.0f}\n")
    else:
        report += "  No renewals within 90 days.\n"

    report += f"\n{'='*60}\nConfidential — CYS Caloyeras · InsureTrack v11.0\n"

    st.markdown(f'<div class="sec-hdr">Executive Summary</div>', unsafe_allow_html=True)
    st.markdown(f"""
    <div style="background:{R['white']};border:1.5px solid {R['rose_lt']};border-radius:12px;
         padding:1.4rem 1.6rem;font-family:'Courier New',monospace;font-size:.78rem;
         color:{R['text_d']};line-height:1.65;max-height:420px;overflow-y:auto;
         white-space:pre-wrap;">{report}</div>
    """, unsafe_allow_html=True)
    st.markdown("<br>", unsafe_allow_html=True)
    c1, c2, c3 = st.columns([1.5, 1.8, 3])
    with c1:
        st.download_button(
            "Download (.txt)",
            data=report,
            file_name=f"insuretrack_summary_{_today().strftime('%Y%m%d')}.txt",
            mime="text/plain",
            key="dl_txt",
        )
    with c2:
        if HAS_FPDF:
            if st.button("Generate PDF", key="gen_pdf"):
                st.session_state["gen_pdf_now"] = True
        else:
            st.caption("PDF: instala `fpdf2`")

    # PDF generation
    if HAS_FPDF and st.session_state.get("gen_pdf_now"):
        st.session_state.pop("gen_pdf_now", None)
        try:
            pdf = FPDF()
            pdf.add_page()
            pdf.set_auto_page_break(auto=True, margin=15)

            # Header
            pdf.set_fill_color(74, 25, 66)   # plum
            pdf.rect(0, 0, 210, 28, "F")
            pdf.set_font("Helvetica", "B", 18)
            pdf.set_text_color(255, 255, 255)
            pdf.set_xy(10, 7)
            pdf.cell(0, 10, "InsureTrack — Portfolio Summary", ln=True)
            pdf.set_font("Helvetica", "", 10)
            pdf.set_xy(10, 19)
            pdf.cell(0, 6, f"{data.get('portfolio_name','CYS Caloyeras')}  ·  {today_str}", ln=True)

            # KPI boxes
            pdf.set_y(35)
            kpi_items = [
                ("Total Properties", str(len(props))),
                ("Insured Active", str(active_n)),
                ("Uninsured", str(uninsu_n)),
                ("Annual Premium", f"${prem_tot:,.0f}"),
                ("Total Units", str(tot_units)),
                ("Auto Premium", f"${auto_prem:,.0f}"),
            ]
            box_w, box_h, gap, cols_n = 55, 20, 5, 3
            start_x = 15
            for i, (lbl, val) in enumerate(kpi_items):
                row = i // cols_n
                col = i % cols_n
                x = start_x + col * (box_w + gap)
                y = 35 + row * (box_h + gap)
                pdf.set_fill_color(253, 246, 240)
                pdf.set_draw_color(242, 196, 206)
                pdf.rect(x, y, box_w, box_h, "FD")
                pdf.set_font("Helvetica", "B", 14)
                pdf.set_text_color(74, 25, 66)
                pdf.set_xy(x + 2, y + 2)
                pdf.cell(box_w - 4, 9, val, align="C")
                pdf.set_font("Helvetica", "", 7)
                pdf.set_text_color(157, 128, 144)
                pdf.set_xy(x + 2, y + 12)
                pdf.cell(box_w - 4, 5, lbl.upper(), align="C")

            y_after_kpis = 35 + 2 * (box_h + gap) + 8

            def section_header(pdf_obj, title, y_pos):
                pdf_obj.set_y(y_pos)
                pdf_obj.set_fill_color(242, 196, 206)
                pdf_obj.rect(10, y_pos, 190, 7, "F")
                pdf_obj.set_font("Helvetica", "B", 9)
                pdf_obj.set_text_color(74, 25, 66)
                pdf_obj.set_xy(12, y_pos + 1)
                pdf_obj.cell(0, 5, title.upper())
                return y_pos + 10

            y = section_header(pdf, "Uninsured Properties", y_after_kpis)
            uninsured_props = sorted([p for p in props if p.get("coverage_status")=="Uninsured"],
                                     key=lambda p: -(p.get("units") or 0))
            pdf.set_font("Helvetica", "", 8)
            pdf.set_text_color(59, 7, 100)
            for p in uninsured_props[:15]:
                pdf.set_x(12)
                pdf.set_y(y)
                txt = f"• {p['prop_id']}  {p.get('address','')}, {p.get('city','')}  ({p.get('units',0)} units)"
                pdf.cell(0, 5, txt[:90], ln=True)
                y += 5
                if y > 270:
                    pdf.add_page(); y = 15

            y = section_header(pdf, "Renewals Next 90 Days", y + 4)
            pdf.set_font("Helvetica", "", 8)
            pdf.set_text_color(59, 7, 100)
            if near:
                for pol, d in near:
                    pdf.set_x(12)
                    pdf.set_y(y)
                    carrier_s = (pol.get("carrier") or "")[:25]
                    txt = f"• {d:3}d  {pol['policy_number']}  {carrier_s}  ${pol.get('premium',0):,.0f}"
                    pdf.cell(0, 5, txt[:90], ln=True)
                    y += 5
                    if y > 270:
                        pdf.add_page(); y = 15
            else:
                pdf.set_x(12); pdf.set_y(y)
                pdf.cell(0, 5, "No renewals within 90 days.", ln=True)

            # Footer
            pdf.set_y(-15)
            pdf.set_font("Helvetica", "I", 7)
            pdf.set_text_color(157, 128, 144)
            pdf.cell(0, 10, f"Confidential — CYS Caloyeras · InsureTrack v9.0 · {today_str}", align="C")

            pdf_bytes = pdf.output()
            st.download_button(
                "Download PDF",
                data=bytes(pdf_bytes),
                file_name=f"insuretrack_portfolio_{_today().strftime('%Y%m%d')}.pdf",
                mime="application/pdf",
                key="dl_pdf_btn",
            )
            st.success("PDF generated. Click the button above to download it.")
        except Exception as ex:
            st.error(f"Error generating PDF: {ex}")

    # ── Excel Export ───────────────────────────────────────────────
    st.markdown(f'<div class="sec-hdr">{phi("microsoft-excel-logo",14)} Export to Excel</div>',
                unsafe_allow_html=True)
    st.caption("Full portfolio workbook: Properties · Policies · Coverage Gaps · Auto")
    today_str = _today().strftime('%Y%m%d')
    xlsx_buf = _build_xlsx(data)
    if xlsx_buf:
        st.download_button(
            f"{phi('download-simple',13)} Download Portfolio (.xlsx)",
            data=xlsx_buf,
            file_name=f"insuretrack_portfolio_{today_str}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_xlsx",
        )
    else:
        st.warning("Install `openpyxl` to enable Excel export.")


# ═══════════════════════════════════════════════════════════════════
#  ADD / EDIT
# ═══════════════════════════════════════════════════════════════════
def page_add(data, save_fn):
    st.markdown(f"""
    <div style="font-size:1.5rem;font-weight:800;color:{R['text_d']};margin-bottom:.2rem;">
      <i class="ph ph-plus-circle" style="font-size:1.4rem;vertical-align:-4px;margin-right:6px;"></i>Add / Edit</div>
    <div style="font-size:.85rem;color:{R['text_m']};margin-bottom:1.2rem;">
      Add new properties, policies, or auto. Edit existing records.</div>
    """, unsafe_allow_html=True)

    st.info(
        "**Note:** Changes made here are saved during this session. "
        "On Streamlit Cloud, the server restarts periodically and JSON edits won't persist. "
        "After saving, use the **⬇️ Download portfolio.json** button below to keep your changes."
    )
    col_dl, _ = st.columns([1.5, 3])
    with col_dl:
        st.download_button(
            "Download portfolio.json",
            data=json.dumps(data, indent=2, ensure_ascii=False),
            file_name="portfolio.json",
            mime="application/json",
            key="dl_portfolio_top",
        )

    tab1, tab2, tab3, tab4, tab5 = st.tabs([
        "Add Property", "Add Policy", "Add Auto",
        "Edit Property", "Edit Policy"
    ])

    # ── TAB 1: ADD PROPERTY ────────────────────────────────────────
    with tab1:
        st.markdown(f'<div class="sec-hdr">New Property</div>', unsafe_allow_html=True)
        st.caption("Fields marked with * are required.")

        st.markdown(f"""<div style="font-size:.72rem;font-weight:700;color:{R['text_m']};
            text-transform:uppercase;letter-spacing:.07em;margin:.8rem 0 .4rem;">
            <i class="ph ph-map-pin" style="font-size:14px;vertical-align:-2px;margin-right:5px;"></i>Location</div>""", unsafe_allow_html=True)
        c1, c2 = st.columns(2)
        with c1:
            np_nick  = st.text_input("Nickname *",
                placeholder="E.g.: Centinela 3632", key="np_nick",
                help="Short name to identify the property in the system")
            np_addr  = st.text_input("Full Address *",
                placeholder="E.g.: 3632 Centinela Ave", key="np_addr")
            np_city  = st.text_input("City *",
                placeholder="E.g.: Los Angeles", key="np_city")
        with c2:
            np_state = st.text_input("State", value="CA", key="np_state")
            np_zip   = st.text_input("ZIP Code", placeholder="E.g.: 90066", key="np_zip")
            np_owner = st.text_input("Owner / LLC",
                placeholder="E.g.: John Caloyeras", key="np_owner")

        st.markdown(f"""<div style="font-size:.72rem;font-weight:700;color:{R['text_m']};
            text-transform:uppercase;letter-spacing:.07em;margin:.8rem 0 .4rem;">
            <i class="ph ph-buildings" style="font-size:14px;vertical-align:-2px;margin-right:5px;"></i>Property Details</div>""", unsafe_allow_html=True)
        d1, d2, d3, d4 = st.columns(4)
        with d1:
            np_type  = st.selectbox("Property Type",
                ["Residential", "Commercial", "Mixed Use", "Vacant Land"], key="np_type")
        with d2:
            np_units = st.number_input("Units", min_value=0, value=1, step=1, key="np_units",
                help="Number of rentable units")
        with d3:
            np_sqft  = st.number_input("Sq Ft", min_value=0, value=0, step=100, key="np_sqft")
        with d4:
            np_yr    = st.number_input("Year Built",
                min_value=1800, max_value=2030, value=2000, step=1, key="np_yr")

        st.markdown(f"""<div style="font-size:.72rem;font-weight:700;color:{R['text_m']};
            text-transform:uppercase;letter-spacing:.07em;margin:.8rem 0 .4rem;">
            <i class="ph ph-shield-check" style="font-size:13px;margin-right:4px;"></i>Insurance Status</div>""", unsafe_allow_html=True)
        s1, s2, s3 = st.columns(3)
        with s1:
            np_cov   = st.selectbox("Current Coverage",
                ["Uninsured","Active","Quote","Verify — Policy on file","External Owner — Verify"],
                key="np_cov",
                help="Current insurance status for this property")
        with s2:
            np_action = st.selectbox("Pending Action",
                ["None","Needs Attention","In Progress","Pending Quote","Resolved"],
                key="np_action",
                help="What needs to be done with this property?")
        with s3:
            np_mort  = st.text_input("Mortgagee", key="np_mort",
                placeholder="Bank or financial institution")
        np_notes = st.text_area("Additional Notes", key="np_notes", height=70,
            placeholder="E.g.: New construction, pending inspection...")

        if st.button("Add Property", key="btn_add_prop", type="primary"):
            if not np_nick or not np_addr or not np_city:
                st.error("Nickname, Address, and City are required.")
            else:
                # Auto-generate next prop ID
                existing_ids = [p["prop_id"] for p in data["properties"]]
                nums = [int(i[1:]) for i in existing_ids if i.startswith("P") and i[1:].isdigit()]
                next_id = f"P{(max(nums)+1):03d}" if nums else "P001"
                new_prop = {
                    "prop_id": next_id,
                    "nickname": np_nick,
                    "address": np_addr,
                    "city": np_city,
                    "state": np_state,
                    "zip": np_zip,
                    "type": np_type,
                    "year_built": int(np_yr) if np_yr else None,
                    "units": int(np_units),
                    "sqft": int(np_sqft) if np_sqft else None,
                    "owner": np_owner,
                    "mortgagee": np_mort,
                    "agent": "",
                    "notes": np_notes,
                    "coverage_status": np_cov,
                    "action_status": np_action if np_action != "None" else "",
                }
                data["properties"].append(new_prop)
                save_fn(data)
                st.success(f"Property {next_id} — {np_nick} added! ({len(data['properties'])} total)")
                st.cache_data.clear()

    # ── TAB 2: ADD POLICY ─────────────────────────────────────────
    with tab2:
        st.markdown(f'<div class="sec-hdr">New Policy</div>', unsafe_allow_html=True)
        prop_opts = {f"{p['prop_id']} — {p['nickname']}": p["prop_id"]
                     for p in data["properties"]}
        c1, c2 = st.columns(2)
        with c1:
            pol_prop    = st.selectbox("Property *", list(prop_opts.keys()), key="pol_prop")
            pol_carrier = st.text_input("Carrier *", key="pol_carrier")
            pol_agency  = st.text_input("Agency / Broker", key="pol_agency")
            pol_num     = st.text_input("Policy Number *", key="pol_num")
            pol_type    = st.selectbox("Policy Type",
                ["Landlord / Dwelling Fire","Commercial Package","BOP",
                 "General Liability","Umbrella","Other"], key="pol_type")
        with c2:
            pol_eff     = st.date_input("Effective Date", key="pol_eff")
            pol_exp     = st.date_input("Expiration Date", key="pol_exp")
            pol_prem    = st.number_input("Annual Premium ($)", min_value=0.0, step=100.0, key="pol_prem")
            pol_bldg    = st.number_input("Building Limit ($)", min_value=0.0, step=1000.0, key="pol_bldg")
            pol_liab    = st.number_input("Liability Limit ($)", min_value=0.0, step=1000.0, key="pol_liab")
            pol_status  = st.selectbox("Status", ["Active","Quote","Expired","Cancelled"], key="pol_status")
        pol_notes = st.text_area("Notes", key="pol_notes", height=70)

        if st.button("Add Policy", key="btn_add_pol", type="primary"):
            if not pol_carrier or not pol_num:
                st.error("Carrier and Policy Number are required.")
            else:
                new_pol = {
                    "prop_id":        prop_opts[pol_prop],
                    "policy_number":  pol_num,
                    "status":         pol_status,
                    "carrier":        pol_carrier,
                    "agency":         pol_agency,
                    "policy_type":    pol_type,
                    "effective_date": str(pol_eff),
                    "expiration_date":str(pol_exp),
                    "premium":        float(pol_prem),
                    "building_limit": float(pol_bldg),
                    "business_income":0,
                    "liability":      float(pol_liab),
                    "ded_aop":        "",
                    "ded_water":      "",
                    "ded_sewer":      "",
                    "inspection":     "",
                    "habitability":   "",
                    "pdf_link":       "",
                    "notes":          pol_notes,
                }
                data["policies"].append(new_pol)
                # Also update property coverage_status to Active if status is Active
                if pol_status == "Active":
                    for p in data["properties"]:
                        if p["prop_id"] == prop_opts[pol_prop]:
                            p["coverage_status"] = "Active"
                elif pol_status == "Quote":
                    for p in data["properties"]:
                        if p["prop_id"] == prop_opts[pol_prop] and p["coverage_status"] == "Uninsured":
                            p["coverage_status"] = "Quote"
                save_fn(data)
                st.success(f"Policy {pol_num} added for {prop_opts[pol_prop]}!")
                st.cache_data.clear()

    # ── TAB 3: ADD AUTO ────────────────────────────────────────────
    with tab3:
        st.markdown(f'<div class="sec-hdr">New Auto Policy</div>', unsafe_allow_html=True)
        c1, c2 = st.columns(2)
        with c1:
            au_carrier = st.text_input("Carrier *", key="au_carrier")
            au_num     = st.text_input("Policy Number *", key="au_num")
            au_driver  = st.text_input("Named Insured / Driver", key="au_driver")
            au_vehicle = st.text_input("Vehicle (Year Make Model)", key="au_vehicle")
        with c2:
            au_eff   = st.date_input("Effective Date", key="au_eff")
            au_exp   = st.date_input("Expiration Date", key="au_exp")
            au_prem  = st.number_input("Annual Premium ($)", min_value=0.0, step=50.0, key="au_prem")
            au_liab  = st.text_input("Liability Limits", placeholder="e.g. 100/300/100", key="au_liab")
        au_notes = st.text_area("Notes", key="au_notes", height=70)

        if st.button("Add Auto Policy", key="btn_add_auto", type="primary"):
            if not au_carrier or not au_num:
                st.error("Carrier and Policy Number are required.")
            else:
                new_auto = {
                    "policy_number":  au_num,
                    "carrier":        au_carrier,
                    "named_insured":  au_driver,
                    "vehicle":        au_vehicle,
                    "effective_date": str(au_eff),
                    "expiration_date":str(au_exp),
                    "premium":        float(au_prem),
                    "liability":      au_liab,
                    "notes":          au_notes,
                }
                if "auto_policies" not in data:
                    data["auto_policies"] = []
                data["auto_policies"].append(new_auto)
                save_fn(data)
                st.success(f"Auto policy {au_num} added!")
                st.cache_data.clear()

    # ── TAB 4: EDIT PROPERTY ───────────────────────────────────────
    with tab4:
        st.markdown(f'<div class="sec-hdr">Edit Property</div>', unsafe_allow_html=True)
        prop_list = {f"{p['prop_id']} — {p['nickname']}": i
                     for i, p in enumerate(data["properties"])}
        if not prop_list:
            st.info("No properties found.")
        else:
            sel_prop = st.selectbox("Select property to edit", list(prop_list.keys()), key="ep_sel")
            idx = prop_list[sel_prop]
            p   = data["properties"][idx]

            c1, c2 = st.columns(2)
            with c1:
                ep_nick  = st.text_input("Nickname", value=p.get("nickname",""), key="ep_nick")
                ep_addr  = st.text_input("Address",  value=p.get("address",""),  key="ep_addr")
                ep_city  = st.text_input("City",     value=p.get("city",""),     key="ep_city")
                ep_state = st.text_input("State",    value=p.get("state",""),    key="ep_state")
                ep_zip   = st.text_input("ZIP",      value=p.get("zip",""),      key="ep_zip")
            with c2:
                type_opts = ["Residential","Commercial","Mixed Use","Vacant Land"]
                ep_type  = st.selectbox("Type", type_opts,
                    index=type_opts.index(p.get("type","Residential")) if p.get("type") in type_opts else 0,
                    key="ep_type")
                ep_units = st.number_input("Units", min_value=0, value=int(p.get("units") or 0), key="ep_units")
                ep_sqft  = st.number_input("Sq Ft", min_value=0, value=int(p.get("sqft") or 0), key="ep_sqft")
                ep_yr    = st.number_input("Year Built", min_value=1800, max_value=2030,
                    value=int(p.get("year_built") or 2000), key="ep_yr")
                ep_owner = st.text_input("Owner / LLC", value=p.get("owner",""), key="ep_owner")
                cov_opts = ["Active","Uninsured","Quote","Verify — Policy on file",
                            "External Owner — Verify","Verify — Policy NN1901886 on file"]
                cur_cov  = p.get("coverage_status","Uninsured")
                if cur_cov not in cov_opts:
                    cov_opts.append(cur_cov)
                ep_cov   = st.selectbox("Coverage Status", cov_opts,
                    index=cov_opts.index(cur_cov), key="ep_cov")
                action_opts = ["None","Needs Attention","In Progress","Pending Quote","Resolved"]
                cur_act  = p.get("action_status","") or "None"
                if cur_act not in action_opts: cur_act = "None"
                ep_action = st.selectbox("Action Status", action_opts,
                    index=action_opts.index(cur_act), key="ep_action")
            ep_mort  = st.text_input("Mortgagee", value=p.get("mortgagee",""), key="ep_mort")
            ep_notes = st.text_area("Notes", value=p.get("notes",""), key="ep_notes", height=70)

            col_sv, col_del, _ = st.columns([1, 1, 4])
            with col_sv:
                if st.button("Save Changes", key="ep_save", type="primary"):
                    data["properties"][idx].update({
                        "nickname":        ep_nick,
                        "address":         ep_addr,
                        "city":            ep_city,
                        "state":           ep_state,
                        "zip":             ep_zip,
                        "type":            ep_type,
                        "units":           int(ep_units),
                        "sqft":            int(ep_sqft) if ep_sqft else None,
                        "year_built":      int(ep_yr) if ep_yr else None,
                        "owner":           ep_owner,
                        "mortgagee":       ep_mort,
                        "notes":           ep_notes,
                        "coverage_status": ep_cov,
                        "action_status":   ep_action if ep_action != "None" else "",
                    })
                    save_fn(data)
                    st.success(f"{ep_nick} updated!")
                    st.cache_data.clear()
            with col_del:
                if st.button("Delete", key="ep_del"):
                    st.session_state["confirm_del_prop"] = idx

            if st.session_state.get("confirm_del_prop") == idx:
                st.warning(f"Delete **{p['prop_id']} — {p.get('nickname','')}**? This cannot be undone.")
                cc1, cc2, _ = st.columns([1,1,4])
                with cc1:
                    if st.button("Yes, delete", key="ep_del_confirm"):
                        data["properties"].pop(idx)
                        save_fn(data)
                        st.session_state.pop("confirm_del_prop", None)
                        st.success("Property deleted.")
                        st.cache_data.clear()
                        st.rerun()
                with cc2:
                    if st.button("Cancel", key="ep_del_cancel"):
                        st.session_state.pop("confirm_del_prop", None)
                        st.rerun()

    # ── TAB 5: EDIT POLICY ─────────────────────────────────────────
    with tab5:
        st.markdown(f'<div class="sec-hdr">Edit Policy</div>', unsafe_allow_html=True)
        pol_list = {f"{pol.get('prop_id','')} · {pol.get('policy_number','')} — {pol.get('carrier','')}": i
                    for i, pol in enumerate(data["policies"])}
        if not pol_list:
            st.info("No policies found.")
        else:
            sel_pol = st.selectbox("Select policy to edit", list(pol_list.keys()), key="epo_sel")
            pidx    = pol_list[sel_pol]
            pol     = data["policies"][pidx]

            c1, c2 = st.columns(2)
            with c1:
                epo_carrier = st.text_input("Carrier", value=pol.get("carrier",""), key="epo_carrier")
                epo_agency  = st.text_input("Agency",  value=pol.get("agency",""),  key="epo_agency")
                epo_num     = st.text_input("Policy Number", value=pol.get("policy_number",""), key="epo_num")
                epo_type    = st.text_input("Policy Type", value=pol.get("policy_type",""), key="epo_type")
            with c2:
                try:
                    _eff_default = (datetime.strptime(pol["effective_date"],  "%Y-%m-%d").date()
                                    if pol.get("effective_date") else _today())
                    _exp_default = (datetime.strptime(pol["expiration_date"], "%Y-%m-%d").date()
                                    if pol.get("expiration_date") else _today())
                except ValueError:
                    _eff_default = _exp_default = _today()
                epo_eff  = st.date_input("Effective Date",  value=_eff_default, key="epo_eff")
                epo_exp  = st.date_input("Expiration Date", value=_exp_default, key="epo_exp")
                epo_prem = st.number_input("Annual Premium ($)",
                    min_value=0.0, value=float(pol.get("premium") or 0), step=100.0, key="epo_prem")
                stat_opts = ["Active","Quote","Expired","Cancelled"]
                cur_stat  = pol.get("status","Active")
                epo_stat  = st.selectbox("Status", stat_opts,
                    index=stat_opts.index(cur_stat) if cur_stat in stat_opts else 0, key="epo_stat")
            epo_notes = st.text_area("Notes", value=pol.get("notes",""), key="epo_notes", height=70)

            col_sv2, col_del2, _ = st.columns([1, 1, 4])
            with col_sv2:
                if st.button("Save Changes", key="epo_save", type="primary"):
                    data["policies"][pidx].update({
                        "carrier":        epo_carrier,
                        "agency":         epo_agency,
                        "policy_number":  epo_num,
                        "policy_type":    epo_type,
                        "effective_date": str(epo_eff),
                        "expiration_date":str(epo_exp),
                        "premium":        float(epo_prem),
                        "status":         epo_stat,
                        "notes":          epo_notes,
                    })
                    save_fn(data)
                    st.success(f"Policy {epo_num} updated!")
                    st.cache_data.clear()
            with col_del2:
                if st.button("Delete", key="epo_del"):
                    st.session_state["confirm_del_pol"] = pidx

            if st.session_state.get("confirm_del_pol") == pidx:
                st.warning(f"Delete policy **{pol.get('policy_number','')}**? This cannot be undone.")
                cc3, cc4, _ = st.columns([1,1,4])
                with cc3:
                    if st.button("Yes, delete", key="epo_del_confirm"):
                        data["policies"].pop(pidx)
                        save_fn(data)
                        st.session_state.pop("confirm_del_pol", None)
                        st.success("Policy deleted.")
                        st.cache_data.clear()
                        st.rerun()
                with cc4:
                    if st.button("Cancel", key="epo_del_cancel"):
                        st.session_state.pop("confirm_del_pol", None)
                        st.rerun()


# ═══════════════════════════════════════════════════════════════════
#  MAP
# ═══════════════════════════════════════════════════════════════════
def page_map(data):
    props = data["properties"]

    st.markdown(f"""
    <div class="page-hdr"><i class="ph ph-map-trifold" style="font-size:24px;vertical-align:-4px;margin-right:8px;"></i>Property Map</div>
    <div class="page-sub">Geographic distribution of the portfolio · {len(props)} properties</div>
    """, unsafe_allow_html=True)

    # Filter controls
    fc1, fc2, fc3 = st.columns([2, 1.5, 1.5])
    with fc1:
        srch = st.text_input("Search city or property", placeholder="e.g. Los Angeles, Santa Monica…",
                             key="map_srch", label_visibility="collapsed")
    with fc2:
        status_opts = ["All"] + sorted(set(p.get("coverage_status","") for p in props if p.get("coverage_status","")))
        flt_status  = st.selectbox("Status", status_opts, key="map_status", label_visibility="collapsed")
    with fc3:
        owner_opts = ["All owners"] + sorted(set(p.get("owner","") for p in props if p.get("owner","")))
        flt_owner  = st.selectbox("Owner", owner_opts, key="map_owner", label_visibility="collapsed")

    filtered = [p for p in props if
        (not srch or srch.lower() in (p.get("city","") or "").lower() or
                     srch.lower() in (p.get("nickname","") or "").lower() or
                     srch.lower() in p.get("prop_id","").lower()) and
        (flt_status == "All" or p.get("coverage_status","") == flt_status) and
        (flt_owner == "All owners" or p.get("owner","") == flt_owner)
    ]

    if not filtered:
        st.info("No properties match the current filter.")
        return

    # Build map data
    lats, lons, texts, colors, sizes = [], [], [], [], []
    color_map = {
        "Active":    "#22C55E",
        "Uninsured": "#EF4444",
        "Quote":     "#F59E0B",
    }
    for p in filtered:
        lat, lon = _get_coords(p.get("city",""), p.get("prop_id",""))
        lats.append(lat); lons.append(lon)
        status = p.get("coverage_status","Unknown")
        units  = p.get("units") or 1
        colors.append(color_map.get(status, "#9D8090"))
        sizes.append(max(10, min(30, 8 + units * 1.5)))
        texts.append(
            f"<b>{p['prop_id']}</b><br>"
            f"{p.get('nickname') or p.get('address','')}<br>"
            f"{p.get('city','')}, {p.get('state','')}<br>"
            f"Status: {status} · {units} units<br>"
            f"Owner: {p.get('owner','—')}"
        )

    fig = go.Figure(go.Scattermapbox(
        lat=lats, lon=lons,
        mode="markers",
        marker=dict(size=sizes, color=colors, opacity=0.85),
        text=texts,
        hovertemplate="%{text}<extra></extra>",
    ))
    fig.update_layout(
        mapbox=dict(style="open-street-map", center=dict(lat=47.5, lon=-122.0), zoom=8),
        height=560,
        margin=dict(l=0, r=0, t=0, b=0),
        paper_bgcolor="rgba(0,0,0,0)",
    )
    st.plotly_chart(fig, use_container_width=True, config={"displayModeBar": True})

    # Legend
    leg_cols = st.columns(4)
    legend = [
        ("#22C55E", "Active"),
        ("#EF4444", "Uninsured"),
        ("#F59E0B", "Quote / Verify"),
        ("#9D8090", "Otro"),
    ]
    for col, (c, lbl) in zip(leg_cols, legend):
        col.markdown(
            f'<div style="display:flex;align-items:center;gap:8px;font-size:.78rem;">'
            f'<div style="width:14px;height:14px;border-radius:50%;background:{c};"></div>'
            f'{lbl}</div>',
            unsafe_allow_html=True
        )

    # Summary table below map
    st.markdown(f'<div class="sec-hdr" style="margin-top:1rem;">Properties on map ({len(filtered)})</div>',
                unsafe_allow_html=True)
    tbl_cols = st.columns([1, 2, 2, 1.2, 1.2, 1.5])
    hdrs = ["ID", "Nickname", "City", "Status", "Units", "Owner"]
    for col, h in zip(tbl_cols, hdrs):
        col.markdown(f"<div style='font-size:.68rem;font-weight:700;color:{R['text_m']};text-transform:uppercase;'>{h}</div>",
                     unsafe_allow_html=True)
    for p in filtered[:30]:
        c1, c2, c3, c4, c5, c6 = st.columns([1, 2, 2, 1.2, 1.2, 1.5])
        status = p.get("coverage_status","")
        c1.markdown(f"<div style='font-size:.78rem;font-weight:700;'>{p['prop_id']}</div>", unsafe_allow_html=True)
        c2.markdown(f"<div style='font-size:.78rem;'>{p.get('nickname') or p.get('address','')[:22]}</div>", unsafe_allow_html=True)
        c3.markdown(f"<div style='font-size:.78rem;'>{p.get('city','')}</div>", unsafe_allow_html=True)
        c4.markdown(coverage_badge(status), unsafe_allow_html=True)
        c5.markdown(f"<div style='font-size:.78rem;'>{p.get('units') or '—'}</div>", unsafe_allow_html=True)
        c6.markdown(f"<div style='font-size:.78rem;color:{R['text_m']};'>{(p.get('owner') or '—')[:18]}</div>", unsafe_allow_html=True)
    if len(filtered) > 30:
        st.caption(f"Showing 30 of {len(filtered)} properties.")


# ═══════════════════════════════════════════════════════════════════
#  TASKS / ACTION ITEMS
# ═══════════════════════════════════════════════════════════════════
TASKS_PATH = os.path.join(os.path.dirname(__file__), "data", "caloyeras", "tasks.json")

def load_tasks():
    if os.path.exists(TASKS_PATH):
        try:
            with open(TASKS_PATH) as f:
                return json.load(f)
        except Exception:
            return []
    return []

def save_tasks(tasks):
    os.makedirs(os.path.dirname(TASKS_PATH), exist_ok=True)
    with open(TASKS_PATH, "w") as f:
        json.dump(tasks, f, indent=2, ensure_ascii=False)

def page_tasks(data, save_fn):
    props    = data["properties"]
    policies = data["policies"]

    st.markdown(f"""
    <div class="page-hdr"><i class="ph ph-check-square" style="font-size:22px;vertical-align:-4px;margin-right:8px;"></i>Task List</div>
    <div class="page-sub">Auto-generated action items + manual tasks for the portfolio</div>
    """, unsafe_allow_html=True)

    # ── Auto-generated action items ────────────────────────────────
    auto_items = []

    # Uninsured properties → high priority
    for p in sorted([x for x in props if x.get("coverage_status") == "Uninsured"],
                    key=lambda x: -(x.get("units") or 0)):
        units = p.get("units") or 0
        pri   = "High" if units >= 10 else ("Medium" if units >= 4 else "Low")
        auto_items.append({
            "type": "auto",
            "priority": pri,
            "icon": "●" if pri == "High" else ("◑" if pri == "Medium" else "○"),
            "title": f"Get insurance for {p['prop_id']} — {p.get('nickname') or p.get('address','')}",
            "detail": f"{p.get('city','')}, {p.get('state','CA')} · {units} units · Owner: {p.get('owner','—')}",
            "category": "Missing Coverage",
        })

    # Expiring policies ≤ 60 days
    for pol in sorted(unique_policies(policies), key=lambda x: x.get("expiration_date","9999")):
        d = _days_to(pol.get("expiration_date"))
        if d is not None and 0 <= d <= 60:
            carrier = (pol.get("carrier") or "").split("/")[0].strip()[:24]
            auto_items.append({
                "type": "auto",
                "priority": "High" if d <= 30 else "Medium",
                "icon": "(!)" if d <= 30 else "·",
                "title": f"Renew policy {pol['policy_number']} — {carrier}",
                "detail": f"Expires in {d} days ({pol.get('expiration_date','')})"
                          f" · ${pol.get('premium',0):,.0f}/yr",
                "category": "Urgent Renewal",
            })

    # Properties with action_status set
    for p in props:
        act = p.get("action_status","").strip()
        if act and act.lower() not in ("", "none", "n/a", "ok", "complete"):
            auto_items.append({
                "type": "auto",
                "priority": "Medium",
                "icon": "·",
                "title": f"{p['prop_id']}: {act}",
                "detail": f"{p.get('nickname') or p.get('address','')} · {p.get('city','')}",
                "category": "Pending Action",
            })

    # ── Manual tasks ────────────────────────────────────────────────
    manual_tasks = load_tasks()

    # Stats
    done_n = sum(1 for t in manual_tasks if t.get("done"))
    total_n = len(auto_items) + len(manual_tasks)
    open_n  = len(auto_items) + (len(manual_tasks) - done_n)

    k1, k2, k3 = st.columns(3)
    for col, (num, lbl, color) in zip([k1, k2, k3], [
        (str(total_n), "Total Items",  R['plum']),
        (str(open_n),  "Pending",      R['rose']),
        (str(done_n),  "Completed",    "#16a34a"),
    ]):
        col.markdown(f"""
        <div style="background:{R['white']};border:1.5px solid {R['rose_lt']};
             border-left:4px solid {color};border-radius:10px;padding:.7rem 1rem;text-align:center;">
          <div style="font-size:1.6rem;font-weight:800;color:{color};">{num}</div>
          <div style="font-size:.65rem;font-weight:700;text-transform:uppercase;
               color:{R['text_m']};letter-spacing:.06em;">{lbl}</div>
        </div>""", unsafe_allow_html=True)

    st.markdown("<div style='height:.6rem'></div>", unsafe_allow_html=True)

    tab_auto, tab_manual = st.tabs(["Auto-Generated", "Manual"])

    # Dismissed auto items stored in session
    dismissed = st.session_state.get("dismissed_auto", set())

    with tab_auto:
        visible = [it for it in auto_items
                   if it["title"] not in dismissed]
        if not visible:
            st.markdown(f"""
            <div style="background:{R['green_lt']};border:1.5px solid #86efac;
                 border-radius:12px;padding:1.2rem 1.5rem;margin-top:.5rem;">
              <span style="font-size:.9rem;color:#14532d;">
                <i class="ph ph-check-circle" style="font-size:15px;color:#16a34a;vertical-align:-2px;"></i> <b>All clear!</b> No pending action items.</span>
            </div>""", unsafe_allow_html=True)
            if dismissed:
                if st.button("↩ Restore dismissed items", key="restore_dismissed"):
                    st.session_state["dismissed_auto"] = set()
                    st.rerun()
        else:
            # Group by category
            by_cat = defaultdict(list)
            for item in visible:
                by_cat[item["category"]].append(item)
            for cat, items in by_cat.items():
                st.markdown(f'<div class="sec-hdr">{cat} ({len(items)})</div>', unsafe_allow_html=True)
                for item in items:
                    pri_color = R['red'] if item['priority']=="High" else (R['amber'] if item['priority']=="Medium" else "#ca8a04")
                    pri_bg    = R['red_lt'] if item['priority']=="High" else (R['amber_lt'] if item['priority']=="Medium" else "#fefce8")
                    col_item, col_btn = st.columns([11, 1])
                    with col_item:
                        st.markdown(f"""
                        <div style="display:flex;align-items:flex-start;gap:10px;
                             padding:.6rem .9rem;background:{pri_bg};
                             border-radius:8px;border-left:3px solid {pri_color};
                             margin-bottom:5px;">
                          <div style="font-size:1.1rem;margin-top:1px;">{item['icon']}</div>
                          <div style="flex:1;">
                            <div style="font-size:.82rem;font-weight:700;color:{R['text_d']};">
                              {item['title']}</div>
                            <div style="font-size:.72rem;color:{R['text_m']};margin-top:2px;">
                              {item['detail']}</div>
                          </div>
                          <div style="font-size:.65rem;font-weight:700;color:{pri_color};
                               text-transform:uppercase;white-space:nowrap;padding-top:3px;">
                            {item['priority']}</div>
                        </div>""", unsafe_allow_html=True)
                    with col_btn:
                        safe_key = "dis_" + str(abs(hash(item['title'])))
                        if st.button("✓", key=safe_key, help="Mark as handled"):
                            dismissed.add(item['title'])
                            st.session_state["dismissed_auto"] = dismissed
                            st.rerun()
            if dismissed:
                st.caption(f"↳ {len(dismissed)} item(s) marked as handled. "
                           "They'll reappear on next session.")

    with tab_manual:
        # Add new task
        st.markdown(f'<div class="sec-hdr"><i class="ph ph-plus-circle" style="font-size:14px;vertical-align:-2px;margin-right:5px;"></i>New Task</div>', unsafe_allow_html=True)
        nc1, nc2, nc3 = st.columns([3, 1.2, 1])
        with nc1:
            new_title = st.text_input("Task description", key="task_title",
                                      label_visibility="collapsed", placeholder="e.g. Call broker for quote...")
        with nc2:
            new_pri = st.selectbox("Priority", ["High","Medium","Low"], key="task_pri",
                                   label_visibility="collapsed")
        with nc3:
            if st.button("Add Task", key="task_add", use_container_width=True):
                if new_title.strip():
                    manual_tasks.append({
                        "id": datetime.now().strftime("%Y%m%d%H%M%S"),
                        "title": new_title.strip(),
                        "priority": new_pri,
                        "done": False,
                        "created": datetime.now().strftime("%Y-%m-%d"),
                    })
                    save_tasks(manual_tasks)
                    st.rerun()

        if not manual_tasks:
            st.info("No manual tasks yet. Add one above.")
        else:
            # Pending first, then done
            st.markdown(f'<div class="sec-hdr">Pending ({len(manual_tasks)-done_n})</div>',
                        unsafe_allow_html=True)
            changed = False
            for i, task in enumerate(manual_tasks):
                if task.get("done"):
                    continue
                pri_color = R['red'] if task['priority']=="High" else (R['amber'] if task['priority']=="Medium" else "#ca8a04")
                t1, t2, t3 = st.columns([0.3, 5, 1])
                with t1:
                    if st.checkbox("", key=f"tck_{task['id']}", value=False):
                        manual_tasks[i]["done"] = True
                        changed = True
                with t2:
                    st.markdown(
                        f"<div style='padding-top:6px;font-size:.83rem;font-weight:600;"
                        f"color:{R['text_d']};'>{task['title']}</div>",
                        unsafe_allow_html=True
                    )
                with t3:
                    st.markdown(
                        f"<div style='padding-top:6px;font-size:.7rem;font-weight:700;"
                        f"color:{pri_color};text-transform:uppercase;'>{task['priority']}</div>",
                        unsafe_allow_html=True
                    )

            if done_n:
                with st.expander(f"Completed ({done_n})", expanded=False):
                    for i, task in enumerate(manual_tasks):
                        if not task.get("done"):
                            continue
                        d1, d2, d3 = st.columns([0.3, 5, 1])
                        with d1:
                            if st.checkbox("", key=f"tck_done_{task['id']}", value=True):
                                pass
                            else:
                                manual_tasks[i]["done"] = False
                                changed = True
                        with d2:
                            st.markdown(
                                f"<div style='padding-top:6px;font-size:.83rem;"
                                f"color:{R['text_m']};text-decoration:line-through;'>{task['title']}</div>",
                                unsafe_allow_html=True
                            )
                        with d3:
                            if st.button("×", key=f"del_task_{task['id']}", help="Delete task"):
                                manual_tasks.pop(i)
                                changed = True
                                break

            if changed:
                save_tasks(manual_tasks)
                st.rerun()


# ═══════════════════════════════════════════════════════════════════
#  ANALYTICS
# ═══════════════════════════════════════════════════════════════════
def page_analytics(data):
    props    = data["properties"]
    policies = data["policies"]

    st.markdown(f"""
    <div class="page-hdr"><i class="ph ph-chart-line" style="font-size:24px;vertical-align:-4px;margin-right:8px;"></i>Analytics</div>
    <div class="page-sub">Cost per unit, owner breakdown, and portfolio metrics</div>
    """, unsafe_allow_html=True)

    tab_cost, tab_owner, tab_carriers, tab_trend = st.tabs([
        f"{phi('currency-dollar',13)} Cost / Unit",
        f"{phi('users',13)} By Owner / LLC",
        f"{phi('buildings',13)} Carrier Risk",
        f"{phi('trend-up',13)} Premium Trend",
    ])

    # Build policy → properties map
    pol_props = defaultdict(list)
    for p in policies:
        if p.get("prop_id"):
            pol_props[p["policy_number"]].append(p["prop_id"])

    # Build prop_id → premium map (split premium across covered props)
    prop_prem = defaultdict(float)
    for pol in unique_policies(policies):
        covered = pol_props.get(pol["policy_number"], [])
        prem    = pol.get("premium") or 0
        if covered:
            split = prem / len(covered)
            for pid in covered:
                prop_prem[pid] += split
        else:
            # No prop linked, skip
            pass

    with tab_cost:
        st.markdown(f'<div class="sec-hdr">Cost per Unit — All Insured Properties</div>',
                    unsafe_allow_html=True)

        rows = []
        for p in props:
            units  = p.get("units") or 0
            prem   = prop_prem.get(p["prop_id"], 0)
            sqft   = p.get("sqft") or 0
            cpu    = prem / units if units > 0 else 0
            cpsf   = prem / sqft  if sqft  > 0 else 0
            rows.append({
                "prop": p,
                "prem": prem,
                "units": units,
                "cpu": cpu,
                "cpsf": cpsf,
                "sqft": sqft,
            })

        # Sort by cpu descending
        rows_sorted = sorted(rows, key=lambda x: -x["cpu"])

        # KPIs
        insured_rows = [r for r in rows_sorted if r["prem"] > 0]
        if insured_rows:
            avg_cpu = sum(r["cpu"] for r in insured_rows) / len(insured_rows)
            max_cpu = max(r["cpu"] for r in insured_rows)
            min_cpu = min(r["cpu"] for r in insured_rows if r["cpu"] > 0)
            k1, k2, k3 = st.columns(3)
            for col, (num, lbl, color) in zip([k1, k2, k3], [
                (f"${avg_cpu:,.0f}", "Avg Cost/Unit",  R['rose']),
                (f"${max_cpu:,.0f}", "Highest CPU",    R['red']),
                (f"${min_cpu:,.0f}", "Lowest CPU",     R['green']),
            ]):
                col.markdown(f"""
                <div style="background:{R['white']};border:1.5px solid {R['rose_lt']};
                     border-left:4px solid {color};border-radius:10px;
                     padding:.7rem 1rem;text-align:center;">
                  <div style="font-size:1.6rem;font-weight:800;color:{color};">{num}</div>
                  <div style="font-size:.65rem;font-weight:700;text-transform:uppercase;
                       color:{R['text_m']};letter-spacing:.06em;">{lbl}</div>
                </div>""", unsafe_allow_html=True)

            st.markdown("<div style='height:.6rem'></div>", unsafe_allow_html=True)

            # Bar chart — CPU by property
            top_rows = insured_rows[:20]
            labels = [r["prop"]["prop_id"] for r in top_rows]
            cpus   = [r["cpu"] for r in top_rows]
            bar_colors = [R['red'] if c > avg_cpu*1.5 else (R['amber'] if c > avg_cpu else R['green'])
                          for c in cpus]
            fig = go.Figure(go.Bar(
                x=labels, y=cpus,
                marker_color=bar_colors,
                text=[f"${c:,.0f}" for c in cpus],
                textposition="outside",
                textfont_size=9,
                hovertemplate="<b>%{x}</b><br>$%{y:,.0f}/unidad<extra></extra>",
            ))
            fig.add_hline(y=avg_cpu, line_dash="dot", line_color=R['plum'],
                          annotation_text=f"Promedio ${avg_cpu:,.0f}",
                          annotation_font_size=10)
            fig.update_layout(
                height=300, margin=dict(l=0, r=0, t=20, b=0),
                paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                xaxis=dict(showgrid=False, tickfont_size=9),
                yaxis=dict(showgrid=True, gridcolor="#F0E8EC",
                           tickformat="$,.0f", tickfont_size=9),
                title=dict(text="Cost per Unit — Top 20 (highest to lowest)",
                           font_size=11, x=0),
            )
            st.plotly_chart(fig, use_container_width=True, config={"displayModeBar": False})

        # Table
        hcols = st.columns([1, 2.2, 1.5, 0.8, 1.2, 1.2, 1.4])
        for col, h in zip(hcols, ["ID","Property","City","Units","Premium/yr","$/unit","$/sqft"]):
            col.markdown(f"<div style='font-size:.67rem;font-weight:700;color:{R['text_m']};text-transform:uppercase;'>{h}</div>",
                         unsafe_allow_html=True)
        for r in rows_sorted:
            c1,c2,c3,c4,c5,c6,c7 = st.columns([1, 2.2, 1.5, 0.8, 1.2, 1.2, 1.4])
            p = r["prop"]
            cpu_color = R['red'] if (r["prem"] > 0 and r["cpu"] > 0 and
                                     r["cpu"] > sum(x["cpu"] for x in insured_rows if x["cpu"]>0)/max(1,len(insured_rows))*1.5) else R['text_d']
            c1.markdown(f"<div style='font-size:.78rem;font-weight:700;'>{p['prop_id']}</div>", unsafe_allow_html=True)
            c2.markdown(f"<div style='font-size:.78rem;'>{p.get('nickname') or p.get('address','')[:22]}</div>", unsafe_allow_html=True)
            c3.markdown(f"<div style='font-size:.78rem;'>{p.get('city','')}</div>", unsafe_allow_html=True)
            c4.markdown(f"<div style='font-size:.78rem;text-align:center;'>{r['units'] or '—'}</div>", unsafe_allow_html=True)
            c5.markdown(f"<div style='font-size:.78rem;'>${r['prem']:,.0f}" if r['prem'] else "<div style='font-size:.78rem;color:#aaa;'>Sin seguro</div>", unsafe_allow_html=True)
            c6.markdown(f"<div style='font-size:.78rem;font-weight:700;color:{cpu_color};'>${r['cpu']:,.0f}" if r['cpu'] else "<div style='font-size:.78rem;color:#aaa;'>—</div>", unsafe_allow_html=True)
            c7.markdown(f"<div style='font-size:.78rem;'>${r['cpsf']:,.2f}" if r['cpsf'] else "<div style='font-size:.78rem;color:#aaa;'>—</div>", unsafe_allow_html=True)

    with tab_owner:
        st.markdown(f'<div class="sec-hdr">Summary by Owner / LLC</div>', unsafe_allow_html=True)

        # Group props by owner
        by_owner = defaultdict(list)
        for p in props:
            owner = p.get("owner") or "Sin asignar"
            by_owner[owner].append(p)

        # Sort by total units descending
        owner_summaries = []
        for owner, owned_props in by_owner.items():
            total_units  = sum(p.get("units") or 0 for p in owned_props)
            total_prem   = sum(prop_prem.get(p["prop_id"],0) for p in owned_props)
            active_n     = sum(1 for p in owned_props if p.get("coverage_status")=="Active")
            uninsu_n     = sum(1 for p in owned_props if p.get("coverage_status")=="Uninsured")
            owner_summaries.append({
                "owner": owner,
                "props": owned_props,
                "units": total_units,
                "prem":  total_prem,
                "active": active_n,
                "uninsu": uninsu_n,
                "n": len(owned_props),
            })
        owner_summaries.sort(key=lambda x: -x["units"])

        # Owner bar chart
        if owner_summaries:
            o_labels = [s["owner"][:18] for s in owner_summaries]
            o_units  = [s["units"] for s in owner_summaries]
            o_prems  = [s["prem"] for s in owner_summaries]
            fig_o = go.Figure()
            fig_o.add_trace(go.Bar(name="Unidades", x=o_labels, y=o_units,
                                   marker_color=R['plum'],
                                   hovertemplate="%{x}<br>%{y} unidades<extra></extra>",
                                   yaxis="y"))
            fig_o.add_trace(go.Bar(name="Prima ($)", x=o_labels, y=o_prems,
                                   marker_color=R['rose'],
                                   hovertemplate="%{x}<br>$%{y:,.0f}<extra></extra>",
                                   yaxis="y2",
                                   opacity=0.7))
            fig_o.update_layout(
                height=280, barmode="group",
                margin=dict(l=0, r=50, t=20, b=0),
                paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                xaxis=dict(showgrid=False, tickfont_size=9),
                yaxis=dict(title="Unidades", showgrid=True, gridcolor="#F0E8EC",
                           tickfont_size=9),
                yaxis2=dict(title="Prima ($)", overlaying="y", side="right",
                            tickformat="$,.0f", tickfont_size=9),
                legend=dict(font_size=10, x=0, y=1.05, orientation="h"),
                title=dict(text="Units and Premium by Owner", font_size=11, x=0),
            )
            st.plotly_chart(fig_o, use_container_width=True, config={"displayModeBar": False})

        for s in owner_summaries:
            warn = f"{s['uninsu']} uninsured" if s['uninsu'] else "Complete"
            warn_color = R['amber'] if s['uninsu'] else R['green']
            with st.expander(
                f"{s['owner']}  ·  {s['n']} props  ·  {s['units']} units  ·  ${s['prem']:,.0f}",
                expanded=False
            ):
                mc1, mc2, mc3, mc4 = st.columns(4)
                mc1.metric("Properties", s['n'])
                mc2.metric("Units",      s['units'])
                mc3.metric("Total Premium", f"${s['prem']:,.0f}")
                mc4.metric("Estado", warn)
                st.markdown("<hr style='border-color:{};margin:.4rem 0;'>".format(R['rose_lt']),
                            unsafe_allow_html=True)
                for p in sorted(s['props'], key=lambda x: -(x.get("units") or 0)):
                    pc1, pc2, pc3, pc4 = st.columns([1.2, 2.5, 1.5, 1.5])
                    pc1.markdown(f"<div style='font-size:.78rem;font-weight:700;'>{p['prop_id']}</div>", unsafe_allow_html=True)
                    pc2.markdown(f"<div style='font-size:.78rem;'>{p.get('nickname') or p.get('address','')[:22]}, {p.get('city','')}</div>", unsafe_allow_html=True)
                    pc3.markdown(coverage_badge(p.get("coverage_status","")), unsafe_allow_html=True)
                    pc4.markdown(f"<div style='font-size:.78rem;'>${prop_prem.get(p['prop_id'],0):,.0f}</div>", unsafe_allow_html=True)

    # ── Tab 3: Carrier Risk ────────────────────────────────────────
    with tab_carriers:
        pols_uniq = data.get("_policies_unique", unique_policies(policies))
        carrier_prem = defaultdict(float)
        for p in pols_uniq:
            carrier_prem[(p.get("carrier") or "Unknown")[:30]] += p.get("premium") or 0
        total_prem = sum(carrier_prem.values()) or 1

        # Warning alerts
        for carrier, prem in sorted(carrier_prem.items(), key=lambda x: -x[1]):
            pct = prem / total_prem * 100
            if pct > 60:
                st.error(f"{phi('warning',14)} High concentration risk: **{carrier}** holds **{pct:.0f}%** of total premium (${prem:,.0f})")
            elif pct > 40:
                st.warning(f"{phi('warning',14)} Concentration alert: **{carrier}** holds **{pct:.0f}%** of total premium (${prem:,.0f})")

        carriers = sorted(carrier_prem.keys(), key=lambda c: -carrier_prem[c])
        prems = [carrier_prem[c] for c in carriers]
        pcts  = [carrier_prem[c] / total_prem * 100 for c in carriers]

        colors_c = ["#D4547A","#4A1942","#F59E0B","#22C55E","#60A5FA","#A78BFA","#FB923C","#34D399"]
        fig_c = go.Figure(go.Bar(
            x=carriers,
            y=prems,
            marker_color=colors_c[:len(carriers)],
            text=[f"${p:,.0f}<br>{pct:.0f}%" for p, pct in zip(prems, pcts)],
            textposition="outside",
        ))
        fig_c.update_layout(
            plot_bgcolor=R['bg'], paper_bgcolor=R['bg'],
            font=dict(color=R['text_d'], size=10),
            margin=dict(l=0,r=0,t=10,b=0), height=300,
            yaxis=dict(tickprefix="$", gridcolor=R['rose_lt']),
            xaxis=dict(gridcolor="rgba(0,0,0,0)"),
        )
        st.plotly_chart(fig_c, use_container_width=True, config={"displayModeBar": False})

        st.markdown(f'<div class="sec-hdr">Carrier Summary</div>', unsafe_allow_html=True)
        for carrier in carriers:
            prem = carrier_prem[carrier]
            pct  = prem / total_prem * 100
            pct_color = "#dc2626" if pct > 60 else "#d97706" if pct > 40 else R['text_d']
            st.markdown(f"""
            <div style="display:flex;justify-content:space-between;padding:.35rem .6rem;
                 font-size:.82rem;border-bottom:1px solid {R['rose_lt']};">
              <span><b>{carrier}</b></span>
              <span>${prem:,.0f}</span>
              <span style="color:{pct_color};font-weight:700;">{pct:.1f}%</span>
            </div>""", unsafe_allow_html=True)

    # ── Tab 4: Premium Trend ───────────────────────────────────────
    with tab_trend:
        _render_premium_trend(data)


# ═══════════════════════════════════════════════════════════════════
#  CALENDAR
# ═══════════════════════════════════════════════════════════════════
def page_calendar(data):
    policies = data["properties"]
    pols     = data["policies"]

    st.markdown(f"""
    <div class="page-hdr"><i class="ph ph-calendar" style="font-size:1.4rem;vertical-align:-4px;margin-right:6px;"></i>Renewal Calendar</div>
    <div class="page-sub">Policy expirations — next 18 months</div>
    """, unsafe_allow_html=True)

    today   = _today()
    # Month picker
    all_months = []
    for delta in range(18):
        m = today.month + delta
        y = today.year + (m - 1) // 12
        m = ((m - 1) % 12) + 1
        all_months.append((y, m))

    # Build events by month
    events_by_month = defaultdict(list)
    for pol in unique_policies(pols):
        exp = pol.get("expiration_date")
        if not exp:
            continue
        try:
            d = datetime.strptime(exp, "%Y-%m-%d").date()
        except Exception:
            continue
        key = (d.year, d.month)
        if key in [m for m in all_months]:
            days_left = (d - today).days
            carrier   = (pol.get("carrier") or "").split("/")[0].strip()[:20]
            events_by_month[key].append({
                "day": d.day,
                "pol": pol["policy_number"],
                "carrier": carrier,
                "prem": pol.get("premium") or 0,
                "days": days_left,
                "date": exp,
            })

    # Controls
    sc1, sc2 = st.columns([1.5, 3])
    with sc1:
        view_ahead = st.selectbox("Show", ["3 months","6 months","12 months","18 months"],
                                  index=1, key="cal_view")
    months_to_show = {"3 months": 3, "6 months": 6, "12 months": 12, "18 months": 18}[view_ahead]

    # Render calendar grid
    MONTH_NAMES = ["Ene","Feb","Mar","Abr","May","Jun","Jul","Ago","Sep","Oct","Nov","Dic"]
    months_subset = all_months[:months_to_show]

    # Show 3 months per row
    for row_start in range(0, len(months_subset), 3):
        row_months = months_subset[row_start:row_start+3]
        cols = st.columns(len(row_months))
        for col, (y, m) in zip(cols, row_months):
            events = sorted(events_by_month.get((y, m), []), key=lambda x: x["day"])
            month_total = sum(e["prem"] for e in events)
            urgent      = [e for e in events if e["days"] <= 60]
            hdr_color   = R['red'] if urgent else (R['amber'] if events else R['green'])
            hdr_bg      = R['red_lt'] if urgent else (R['amber_lt'] if events else R['green_lt'])

            with col:
                st.markdown(f"""
                <div style="border:1.5px solid {hdr_color};border-radius:12px;overflow:hidden;
                     margin-bottom:8px;">
                  <div style="background:{hdr_bg};padding:.5rem .8rem;
                       border-bottom:1.5px solid {hdr_color};">
                    <div style="font-weight:800;font-size:.88rem;color:{hdr_color};">
                      {MONTH_NAMES[m-1]} {y}</div>
                    <div style="font-size:.7rem;color:{R['text_m']};">
                      {len(events)} {'policy' if len(events)==1 else 'policies'}{f' · ${month_total:,.0f}' if month_total else ''}</div>
                  </div>
                  <div style="background:{R['white']};padding:.5rem .8rem;">
                """, unsafe_allow_html=True)

                if events:
                    for e in events:
                        dot_color = R['red'] if e['days'] <= 30 else (R['amber'] if e['days'] <= 60 else R['plum'])
                        st.markdown(f"""
                        <div style="display:flex;align-items:center;gap:6px;
                             padding:3px 0;border-bottom:1px solid {R['rose_lt']};">
                          <div style="width:28px;font-size:.72rem;font-weight:800;
                               color:{dot_color};text-align:center;">{e['day']}</div>
                          <div style="flex:1;overflow:hidden;">
                            <div style="font-size:.72rem;font-weight:700;color:{R['text_d']};
                                 white-space:nowrap;overflow:hidden;text-overflow:ellipsis;">
                              {e['pol']}</div>
                            <div style="font-size:.65rem;color:{R['text_m']};">
                              {e['carrier']} · ${e['prem']:,.0f}</div>
                          </div>
                          <div style="font-size:.65rem;color:{dot_color};font-weight:700;
                               white-space:nowrap;">{e['days']}d</div>
                        </div>""", unsafe_allow_html=True)
                else:
                    st.markdown(f"<div style='font-size:.75rem;color:{R['text_m']};padding:.3rem 0;'>No expirations</div>",
                                unsafe_allow_html=True)

                st.markdown("</div></div>", unsafe_allow_html=True)

    # Summary table
    st.markdown(f'<div class="sec-hdr" style="margin-top:1rem;">All Renewals — Next {months_to_show} Months</div>',
                unsafe_allow_html=True)
    all_events = []
    for (y, m) in months_subset:
        for e in events_by_month.get((y, m), []):
            all_events.append(e)
    all_events.sort(key=lambda x: x["date"])

    if all_events:
        sc1, sc2, sc3, sc4 = st.columns([2, 2.5, 1.5, 1.2])
        for col, h in zip([sc1,sc2,sc3,sc4], ["Date","Policy / Carrier","Premium","Days"]):
            col.markdown(f"<div style='font-size:.67rem;font-weight:700;color:{R['text_m']};text-transform:uppercase;'>{h}</div>",
                         unsafe_allow_html=True)
        for e in all_events:
            c1,c2,c3,c4 = st.columns([2, 2.5, 1.5, 1.2])
            d_color = R['red'] if e['days']<=30 else (R['amber'] if e['days']<=60 else R['text_d'])
            c1.markdown(f"<div style='font-size:.78rem;color:{d_color};font-weight:600;'>{e['date']}</div>", unsafe_allow_html=True)
            c2.markdown(f"<div style='font-size:.78rem;'>{e['pol']} · {e['carrier']}</div>", unsafe_allow_html=True)
            c3.markdown(f"<div style='font-size:.78rem;'>${e['prem']:,.0f}</div>", unsafe_allow_html=True)
            c4.markdown(f"<div style='font-size:.78rem;font-weight:700;color:{d_color};'>{e['days']}d</div>", unsafe_allow_html=True)
    else:
        st.info(f"No renewals in the next {months_to_show} months.")


# ═══════════════════════════════════════════════════════════════════
#  IMPORT CSV / EXCEL
# ═══════════════════════════════════════════════════════════════════
def page_import(data, save_fn):
    st.markdown(f"""
    <div class="page-hdr"><i class="ph ph-upload-simple" style="font-size:24px;vertical-align:-4px;margin-right:8px;"></i>Import Data</div>
    <div class="page-sub">Load properties or policies from CSV or Excel (.xlsx)</div>
    """, unsafe_allow_html=True)

    if not HAS_PANDAS:
        st.error("`pandas` is not installed. Add `pandas>=2.0.0` to requirements.txt and restart.")
        return

    tab_props, tab_pols = st.tabs(["Properties", "Policies"])

    PROP_COLS = ["prop_id","nickname","address","city","state","zip","type",
                 "units","sqft","year_built","owner","mortgagee","coverage_status","action_status","notes"]
    POL_COLS  = ["prop_id","policy_number","carrier","agency","policy_type",
                 "effective_date","expiration_date","premium","status","notes"]

    def render_import_tab(entity, required_cols, all_cols, current_list, label):
        st.markdown(f"""
        <div style="background:{R['amber_lt']};border:1.5px solid {R['amber']};
             border-radius:10px;padding:.8rem 1.2rem;margin-bottom:.8rem;font-size:.82rem;color:#78350f;">
          <b>Required columns:</b> {', '.join(f'<code>{c}</code>' for c in required_cols)}<br>
          <b>All supported columns:</b> {', '.join(f'<code>{c}</code>' for c in all_cols)}
        </div>""", unsafe_allow_html=True)

        uploaded = st.file_uploader(
            f"Upload file for {label}",
            type=["csv","xlsx","xls"],
            key=f"upld_{entity}",
            label_visibility="collapsed",
        )
        if uploaded is None:
            st.info(f"Drag or select a CSV/Excel file with the indicated columns.")

            # Template download
            import_template = ",".join(all_cols) + "\n"
            st.download_button(
                f"⬇️ Download CSV template for {label}",
                data=import_template,
                file_name=f"insuretrack_{entity}_template.csv",
                mime="text/csv",
                key=f"tmpl_{entity}",
            )
            return

        try:
            if uploaded.name.endswith(".csv"):
                df = pd.read_csv(uploaded)
            else:
                df = pd.read_excel(uploaded)
        except Exception as ex:
            st.error(f"Error reading file: {ex}")
            return

        st.markdown(f'<div class="sec-hdr">Preview — {len(df)} rows, {len(df.columns)} columns</div>',
                    unsafe_allow_html=True)
        st.dataframe(df.head(20), use_container_width=True)

        # Validate required columns
        missing = [c for c in required_cols if c not in df.columns]
        if missing:
            st.error(f"Missing columns: {', '.join(missing)}")
            return

        # Clean and convert
        df = df.where(pd.notnull(df), None)

        # Check for duplicates
        existing_ids = set(r.get(required_cols[0],"") for r in current_list)
        new_rows   = df[~df[required_cols[0]].astype(str).isin(existing_ids)]
        dup_rows   = df[df[required_cols[0]].astype(str).isin(existing_ids)]

        st.markdown(f"""
        <div style="display:flex;gap:12px;margin:.5rem 0;">
          <div style="background:{R['green_lt']};border:1px solid #86efac;border-radius:8px;
               padding:.5rem .9rem;font-size:.82rem;color:#14532d;">
            <b>{len(new_rows)}</b> new {label}
          </div>
          <div style="background:{R['amber_lt']};border:1px solid #fcd34d;border-radius:8px;
               padding:.5rem .9rem;font-size:.82rem;color:#78350f;">
            <b>{len(dup_rows)}</b> already exist (will be skipped)
          </div>
        </div>""", unsafe_allow_html=True)

        if len(new_rows) == 0:
            st.warning("No new rows to import.")
            return

        if st.button(f"Confirm import of {len(new_rows)} {label}", key=f"confirm_{entity}",
                     type="primary", use_container_width=False):
            added = 0
            for _, row in new_rows.iterrows():
                rec = {col: (row[col] if col in row and row[col] is not None else "")
                       for col in all_cols}
                # Type coercions
                for int_col in ["units","sqft","year_built"]:
                    if int_col in rec and rec[int_col] not in (None,""):
                        try: rec[int_col] = int(rec[int_col])
                        except Exception: rec[int_col] = None
                for flt_col in ["premium"]:
                    if flt_col in rec and rec[flt_col] not in (None,""):
                        try: rec[flt_col] = float(rec[flt_col])
                        except Exception: rec[flt_col] = 0.0
                current_list.append(rec)
                added += 1

            if entity == "props":
                data["properties"] = current_list
            else:
                data["policies"] = current_list

            save_fn(data)
            log_change("import", f"Imported {added} {label} from {uploaded.name}")
            st.success(f"{added} {label} imported successfully.")
            st.cache_data.clear()
            st.rerun()

    with tab_props:
        render_import_tab(
            "props",
            required_cols=["prop_id","address","city"],
            all_cols=PROP_COLS,
            current_list=data["properties"],
            label="properties",
        )
    with tab_pols:
        render_import_tab(
            "pols",
            required_cols=["prop_id","policy_number"],
            all_cols=POL_COLS,
            current_list=data["policies"],
            label="policies",
        )


# ═══════════════════════════════════════════════════════════════════
#  SETTINGS
# ═══════════════════════════════════════════════════════════════════
def page_settings(data, save_fn):
    st.markdown(f"""
    <div style="font-size:1.5rem;font-weight:800;color:{R['text_d']};margin-bottom:.2rem;">
      <i class="ph ph-gear" style="font-size:24px;vertical-align:-4px;margin-right:8px;"></i>Settings</div>
    <div style="font-size:.85rem;color:{R['text_m']};margin-bottom:1.2rem;">
      Portfolio configuration</div>
    """, unsafe_allow_html=True)

    st.markdown(f'<div class="sec-hdr">Portfolio</div>', unsafe_allow_html=True)
    col, _ = st.columns([1.4, 2])
    with col:
        new_name = st.text_input("Portfolio name", value=data.get("portfolio_name",""),
                                 key="sett_nm")
        new_date = st.text_input("As of date (YYYY-MM-DD)",
                                 value=data.get("as_of_date",""), key="sett_dt")
        if st.button("Save Changes"):
            data["portfolio_name"] = new_name
            data["as_of_date"]     = new_date
            save_fn(data)
            st.success("✓ Saved.")

    st.markdown(f'<div class="sec-hdr">Account</div>', unsafe_allow_html=True)
    col2, _ = st.columns([1, 3])
    with col2:
        st.markdown(f"Signed in as **{st.session_state.get('username','')}**")
        if st.button("Sign Out"):
            for k in list(st.session_state.keys()):
                del st.session_state[k]
            st.rerun()

    st.markdown(f'<div class="sec-hdr">Data</div>', unsafe_allow_html=True)
    st.markdown(f"""
    <div class="rg-card-sm" style="font-size:.82rem;color:{R['text_m']};">
      <b>Source:</b> portfolio.json<br>
      <b>Properties:</b> {len(data['properties'])} &nbsp;·&nbsp;
      <b>Policies:</b> {len(unique_policies(data['policies']))} unique &nbsp;·&nbsp;
      <b>Auto:</b> {len(data.get('auto_policies',[]))}
    </div>
    """, unsafe_allow_html=True)

    # ── Email alert config ─────────────────────────────────────────
    st.markdown(f'<div class="sec-hdr"><i class="ph ph-envelope" style="font-size:14px;vertical-align:-2px;margin-right:5px;"></i>Email Alerts</div>', unsafe_allow_html=True)
    st.markdown(f"""
    <div style="background:{R['amber_lt']};border:1.5px solid {R['amber']};
         border-radius:10px;padding:.8rem 1.2rem;font-size:.82rem;color:#78350f;margin-bottom:.8rem;">
      To enable automatic alerts, configure the following variables in
      <code>st.secrets</code> (file <code>.streamlit/secrets.toml</code>):<br><br>
      <code>[email]<br>
      smtp_host = "smtp.gmail.com"<br>
      smtp_port = 587<br>
      smtp_user = "you@email.com"<br>
      smtp_password = "app_password"<br>
      alert_recipient = "caloyeras@email.com"<br>
      days_before_alert = 60</code>
    </div>
    """, unsafe_allow_html=True)

    # Show current email config status
    try:
        email_cfg = st.secrets.get("email", {})
        if email_cfg.get("smtp_host"):
            st.success(f"Email configured: {email_cfg.get('smtp_user','—')} → {email_cfg.get('alert_recipient','—')}")
        else:
            st.info("Email not configured. Use the code block above in secrets.toml.")
    except Exception:
        st.info("Email not configured. Use the code block above in secrets.toml.")

    col_email, _ = st.columns([1.5, 2])
    with col_email:
        if st.button("Send summary now (demo)", key="send_email_demo"):
            st.info("This feature requires SMTP configuration in secrets.toml")

    # ── Changelog ─────────────────────────────────────────────────
    st.markdown(f'<div class="sec-hdr"><i class="ph ph-clock-counter-clockwise" style="font-size:14px;vertical-align:-2px;margin-right:5px;"></i>Change History</div>', unsafe_allow_html=True)
    changelog = load_changelog()
    if not changelog:
        st.markdown(f"""
        <div class="rg-card-sm" style="font-size:.82rem;color:{R['text_m']};">
          No changes recorded yet. Changes are logged automatically when
          importing data or modifying the portfolio.
        </div>""", unsafe_allow_html=True)
    else:
        # Filter controls
        chg_col1, chg_col2, _ = st.columns([1.5, 1.5, 3])
        with chg_col1:
            chg_limit = st.selectbox("Show", ["Last 25","Last 50","All"], key="chg_limit")
        limit_n = {"Last 25": 25, "Last 50": 50, "All": len(changelog)}[chg_limit]

        for entry in changelog[:limit_n]:
            ts   = entry.get("timestamp","—")
            user = entry.get("user","—")
            act  = entry.get("action","—")
            det  = entry.get("detail","")
            act_icons = {"import":"↑","save":"·","delete":"×","edit":"✎"}
            icon = act_icons.get(act, "·")
            st.markdown(f"""
            <div style="display:flex;gap:10px;padding:.45rem .7rem;
                 background:{R['white']};border:1px solid {R['rose_lt']};
                 border-radius:8px;margin-bottom:4px;align-items:flex-start;">
              <div style="font-size:1rem;min-width:24px;">{icon}</div>
              <div style="flex:1;">
                <div style="font-size:.78rem;font-weight:700;color:{R['text_d']};">
                  {act.title()} — {det[:80]}</div>
                <div style="font-size:.67rem;color:{R['text_m']};">
                  {ts} · {user}</div>
              </div>
            </div>""", unsafe_allow_html=True)

        if len(changelog) > limit_n:
            st.caption(f"Showing {limit_n} of {len(changelog)} entries.")

    # ── Alert Email Preview ────────────────────────────────────────
    st.markdown(f'<div class="sec-hdr">{phi("envelope",14)} Alert Email Preview</div>',
                unsafe_allow_html=True)
    st.caption("Preview of what the alert email would say if sent today.")

    props = data["properties"]
    pols  = data["policies"]
    today_label = _today().strftime("%B %d, %Y")

    expiring_soon = sorted(
        [(pol, _days_to(pol.get("expiration_date"))) for pol in pols
         if _days_to(pol.get("expiration_date")) is not None and 0 <= _days_to(pol.get("expiration_date")) <= 60],
        key=lambda x: x[1]
    )
    uninsured = [p for p in props if p.get("coverage_status") in ("Uninsured", "Lapsed")]

    exp_lines = "\n".join(
        f"  • {p.get('policy_number','')} – {(p.get('carrier') or '')[:25]} – {d}d"
        for p, d in expiring_soon[:15]
    ) or "  None within 60 days."
    uni_lines = "\n".join(
        f"  • {p.get('prop_id','')} – {p.get('nickname') or p.get('address','')} – {p.get('coverage_status','')}"
        for p in uninsured[:15]
    ) or "  None."

    preview = (
        f"Subject: InsureTrack Alert — {len(expiring_soon)} renewals due, {len(uninsured)} uninsured\n"
        f"Date: {today_label}\n"
        f"{'─'*52}\n\n"
        f"POLICIES EXPIRING WITHIN 60 DAYS ({len(expiring_soon)}):\n{exp_lines}\n\n"
        f"UNINSURED / LAPSED PROPERTIES ({len(uninsured)}):\n{uni_lines}\n\n"
        f"Portfolio: {data.get('portfolio_name','CYS Caloyeras')} · InsureTrack v11.0\n"
    )

    st.markdown(f"""
    <div style="background:{R['white']};border:1.5px solid {R['rose_lt']};border-radius:12px;
         padding:1.2rem 1.4rem;font-family:'Courier New',monospace;font-size:.75rem;
         color:{R['text_d']};line-height:1.65;max-height:320px;overflow-y:auto;
         white-space:pre-wrap;">{preview}</div>
    """, unsafe_allow_html=True)
    c_email, _ = st.columns([1.8, 3])
    with c_email:
        if st.button(f"{phi('paper-plane-tilt',13)} Send Test Email (demo)", key="send_test_email"):
            st.info("SMTP not configured — add credentials to secrets.toml to enable.")


# ═══════════════════════════════════════════════════════════════════
#  GLOBAL SEARCH
# ═══════════════════════════════════════════════════════════════════
def page_search(data):
    st.markdown(f"""
    <div class="page-hdr">{phi("magnifying-glass", 24)} Search</div>
    <div class="page-sub">Find any property, policy, carrier or owner across the portfolio</div>
    """, unsafe_allow_html=True)

    q = st.text_input("", placeholder="Type to search…", key="search_main",
                      value=st.session_state.get("search_query", ""),
                      label_visibility="collapsed")
    st.session_state["search_query"] = q

    if not q or len(q) < 2:
        st.info("Enter at least 2 characters to search.")
        return

    ql = q.lower()
    props = data["properties"]
    pols  = data.get("_policies_unique", unique_policies(data["policies"]))
    autos = data.get("auto_policies", [])

    prop_hits = [p for p in props if
                 ql in p.get("address","").lower() or
                 ql in p.get("nickname","").lower() or
                 ql in p.get("city","").lower() or
                 ql in (p.get("owner") or "").lower() or
                 ql in p.get("prop_id","").lower()]

    pol_hits  = [p for p in pols if
                 ql in p.get("policy_number","").lower() or
                 ql in (p.get("carrier") or "").lower() or
                 ql in (p.get("agency") or "").lower() or
                 ql in p.get("prop_id","").lower()]

    auto_hits = [a for a in autos if
                 ql in (a.get("carrier") or "").lower() or
                 ql in (a.get("policy_number") or "").lower() or
                 ql in (a.get("vehicle") or "").lower()]

    total = len(prop_hits) + len(pol_hits) + len(auto_hits)
    st.markdown(f'<div class="sec-hdr">{phi("funnel",14)} {total} results for "{q}"</div>',
                unsafe_allow_html=True)

    if prop_hits:
        st.markdown(f'<div style="font-size:.78rem;font-weight:700;color:{R["text_m"]};margin:.8rem 0 .4rem;">'
                    f'{phi("buildings",13)} Properties ({len(prop_hits)})</div>', unsafe_allow_html=True)
        for p in prop_hits[:20]:
            col1, col2 = st.columns([5, 1])
            with col1:
                st.markdown(f"""
                <div class="rg-card-sm" style="margin-bottom:4px;">
                  <b>{p.get("nickname") or p["prop_id"]}</b>
                  &nbsp;·&nbsp; {p.get("address","")}, {p.get("city","")}
                  &nbsp; {coverage_badge(p.get("coverage_status",""))}
                </div>""", unsafe_allow_html=True)
            with col2:
                if st.button("View", key=f"sr_p_{p['prop_id']}"):
                    st.session_state.page = "properties"
                    st.rerun()

    if pol_hits:
        st.markdown(f'<div style="font-size:.78rem;font-weight:700;color:{R["text_m"]};margin:.8rem 0 .4rem;">'
                    f'{phi("file-text",13)} Policies ({len(pol_hits)})</div>', unsafe_allow_html=True)
        for p in pol_hits[:20]:
            d = _days_to(p.get("expiration_date"))
            urgency = (f'<span style="color:#dc2626;font-size:.72rem;">{d}d</span>' if d is not None and d < 60 else
                       f'<span style="font-size:.72rem;color:{R["text_m"]};">{d}d</span>' if d is not None else "")
            col1, col2 = st.columns([5, 1])
            with col1:
                st.markdown(f"""
                <div class="rg-card-sm" style="margin-bottom:4px;">
                  <b>{p.get("policy_number","")}</b>
                  &nbsp;·&nbsp; {(p.get("carrier") or "")[:30]}
                  &nbsp;·&nbsp; {p.get("prop_id","")} {urgency}
                </div>""", unsafe_allow_html=True)
            with col2:
                if st.button("View", key=f"sr_pol_{p.get('policy_number','')}"):
                    st.session_state.page = "policies"
                    st.rerun()

    if auto_hits:
        st.markdown(f'<div style="font-size:.78rem;font-weight:700;color:{R["text_m"]};margin:.8rem 0 .4rem;">'
                    f'{phi("car",13)} Auto Policies ({len(auto_hits)})</div>', unsafe_allow_html=True)
        for a in auto_hits[:10]:
            st.markdown(f"""
            <div class="rg-card-sm" style="margin-bottom:4px;">
              <b>{a.get("policy_number","")}</b>
              &nbsp;·&nbsp; {(a.get("carrier") or "")[:30]}
              &nbsp;·&nbsp; {a.get("vehicle","")}
            </div>""", unsafe_allow_html=True)

    if total == 0:
        st.warning(f"No results found for '{q}'.")


# ═══════════════════════════════════════════════════════════════════
#  RENEWAL ACTION QUEUE
# ═══════════════════════════════════════════════════════════════════
def page_queue(data):
    st.markdown(f"""
    <div class="page-hdr">{phi("clock-countdown", 24)} Renewal Action Queue</div>
    <div class="page-sub">Policies requiring action in the next 90 days</div>
    """, unsafe_allow_html=True)

    props   = data["properties"]
    pols    = data["policies"]
    prop_by_id = {p["prop_id"]: p for p in props}

    snoozed = st.session_state.get("snoozed_queue", set())

    # Build queue items
    queue = []
    today = _today()
    for pol in pols:
        if pol.get("policy_number") in snoozed:
            continue
        d = _days_to(pol.get("expiration_date"))
        if d is None:
            continue
        if d <= 90:
            prop = prop_by_id.get(pol.get("prop_id",""), {})
            queue.append((d, pol, prop))
    queue.sort(key=lambda x: x[0])

    overdue_n = sum(1 for d,_,__ in queue if d < 0)
    lt30_n    = sum(1 for d,_,__ in queue if 0 <= d < 30)
    lt60_n    = sum(1 for d,_,__ in queue if 30 <= d < 60)
    lt90_n    = sum(1 for d,_,__ in queue if 60 <= d <= 90)

    # Summary bar
    st.markdown(f"""
    <div style="display:flex;gap:12px;padding:.5rem .8rem;background:{R['white']};
         border:1.5px solid {R['rose_lt']};border-radius:8px;margin-bottom:1rem;font-size:.78rem;">
      <span><b>{len(queue)}</b> policies</span>
      <span style="color:#dc2626;">{phi("warning",13,color="#dc2626")} <b>{overdue_n}</b> overdue</span>
      <span style="color:#ea580c;"><b>{lt30_n}</b> within 30d</span>
      <span style="color:#d97706;"><b>{lt60_n}</b> within 60d</span>
      <span style="color:{R['text_m']};"><b>{lt90_n}</b> within 90d</span>
    </div>
    """, unsafe_allow_html=True)

    # Filter
    filt_opts = ["All", "Overdue", "< 30 days", "31–60 days", "61–90 days"]
    filt = st.radio("", filt_opts, horizontal=True, key="queue_filt", label_visibility="collapsed")

    def in_filt(d):
        if filt == "All":         return True
        if filt == "Overdue":     return d < 0
        if filt == "< 30 days":   return 0 <= d < 30
        if filt == "31–60 days":  return 30 <= d < 60
        if filt == "61–90 days":  return 60 <= d <= 90
        return True

    visible = [(d, pol, prop) for d, pol, prop in queue if in_filt(d)]

    if not visible:
        st.success(f"No policies in this filter range.")
        return

    for d, pol, prop in visible:
        if d < 0:
            badge_color, badge_bg = "#991b1b", "#fee2e2"
            badge_txt = f"Overdue {abs(d)}d"
        elif d < 30:
            badge_color, badge_bg = "#9a3412", "#ffedd5"
            badge_txt = f"{d}d left"
        elif d < 60:
            badge_color, badge_bg = "#92400e", "#fef3c7"
            badge_txt = f"{d}d left"
        else:
            badge_color, badge_bg = R['text_m'], R['gray_lt']
            badge_txt = f"{d}d left"

        pnum = pol.get("policy_number","—")
        with st.container():
            st.markdown(f"""
            <div style="background:{R['white']};border:1.5px solid {R['rose_lt']};
                 border-radius:10px;padding:.9rem 1.1rem;margin-bottom:.5rem;">
              <div style="display:flex;justify-content:space-between;align-items:flex-start;">
                <div>
                  <div style="font-size:.9rem;font-weight:700;color:{R['text_d']};">
                    {prop.get('nickname') or prop.get('prop_id','—')}
                    &nbsp;<span style="font-weight:400;font-size:.78rem;color:{R['text_m']};">{prop.get('address','')}, {prop.get('city','')}</span>
                  </div>
                  <div style="font-size:.78rem;color:{R['text_m']};margin-top:3px;">
                    {phi("file-text",12)} {pnum} &nbsp;·&nbsp; {(pol.get('carrier') or 'No carrier')[:35]}
                    &nbsp;·&nbsp; ${pol.get('premium') or 0:,.0f}
                  </div>
                </div>
                <span style="background:{badge_bg};color:{badge_color};border-radius:20px;
                      padding:3px 12px;font-size:.75rem;font-weight:700;white-space:nowrap;">{badge_txt}</span>
              </div>
            </div>""", unsafe_allow_html=True)

            c1, c2, c3 = st.columns([1.5, 1.5, 1.5])
            with c1:
                if st.button(f"{phi('check-circle',12)} Mark Renewed", key=f"qmr_{pnum}"):
                    log_change("renew", f"Policy {pnum} marked renewed from queue")
                    st.success(f"Marked {pnum} as renewed.")
            with c2:
                if st.button(f"{phi('chat-circle-text',12)} Request Quote", key=f"qrq_{pnum}"):
                    st.info(f"Flagged {pnum} for quote request.")
            with c3:
                if st.button(f"{phi('clock',12)} Snooze 30d", key=f"qsn_{pnum}"):
                    snoozed.add(pnum)
                    st.session_state["snoozed_queue"] = snoozed
                    st.rerun()

    if snoozed:
        st.caption(f"{len(snoozed)} snoozed this session.")
        if st.button("Clear all snoozes", key="clear_snooze"):
            st.session_state["snoozed_queue"] = set()
            st.rerun()


# ═══════════════════════════════════════════════════════════════════
#  XLSX EXPORT (helper — used from Reports page)
# ═══════════════════════════════════════════════════════════════════
def _build_xlsx(data):
    """Build a formatted portfolio workbook and return BytesIO."""
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return None

    wb = openpyxl.Workbook()

    HDR_FILL  = PatternFill("solid", fgColor="4A1942")
    HDR_FONT  = Font(color="FFFFFF", bold=True, size=10)
    ACT_FILL  = PatternFill("solid", fgColor="F0FDF4")
    UNI_FILL  = PatternFill("solid", fgColor="FEE2E2")
    AMB_FILL  = PatternFill("solid", fgColor="FFFBEB")
    THIN_SIDE = Side(style="thin", color="E5E7EB")
    THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)

    def _write_sheet(ws, headers, rows, row_fill_fn=None):
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = HDR_FILL
            cell.font = HDR_FONT
            cell.alignment = Alignment(horizontal="left")
        ws.freeze_panes = "A2"
        for ri, row in enumerate(rows, start=2):
            ws.append(row)
            if row_fill_fn:
                fill = row_fill_fn(row)
                if fill:
                    for cell in ws[ri]:
                        cell.fill = fill
        for col_cells in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col_cells), default=8)
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 2, 40)

    props  = data["properties"]
    pols   = data["policies"]
    autos  = data.get("auto_policies", [])
    pol_by_prop = {}
    for p in pols:
        pol_by_prop.setdefault(p.get("prop_id",""), []).append(p)

    # Sheet 1 — Properties
    ws1 = wb.active
    ws1.title = "Properties"
    prop_hdrs = ["prop_id","nickname","address","city","state","zip","type",
                 "units","sqft","year_built","owner","coverage_status","action_status","notes"]
    prop_rows = [[p.get(h,"") for h in prop_hdrs] for p in props]
    def prop_fill(row):
        s = row[11] if len(row) > 11 else ""
        if s == "Active":     return ACT_FILL
        if s == "Uninsured":  return UNI_FILL
        if s:                 return AMB_FILL
    _write_sheet(ws1, prop_hdrs, prop_rows, prop_fill)

    # Sheet 2 — Policies
    ws2 = wb.create_sheet("Policies")
    pol_hdrs = ["policy_number","prop_id","carrier","agency","policy_type",
                "effective_date","expiration_date","premium","status","notes"]
    pol_rows = [[p.get(h,"") for h in pol_hdrs] for p in pols]
    def pol_fill(row):
        exp = row[6] if len(row) > 6 else ""
        d = _days_to(str(exp)) if exp else None
        if d is not None and d < 0:   return UNI_FILL
        if d is not None and d < 60:  return AMB_FILL
    _write_sheet(ws2, pol_hdrs, pol_rows, pol_fill)

    # Sheet 3 — Coverage Gaps
    ws3 = wb.create_sheet("Coverage Gaps")
    gap_hdrs = ["prop_id","nickname","address","city","owner","units","coverage_status","action_status","notes"]
    gap_props = [p for p in props if p.get("coverage_status") not in ("Active",)]
    gap_rows  = [[p.get(h,"") for h in gap_hdrs] for p in gap_props]
    _write_sheet(ws3, gap_hdrs, gap_rows, lambda r: UNI_FILL if r[6]=="Uninsured" else AMB_FILL)

    # Sheet 4 — Auto
    ws4 = wb.create_sheet("Auto")
    auto_hdrs = ["policy_number","carrier","vehicle","state","effective_date",
                 "expiration_date","premium","status","notes"]
    auto_rows = [[a.get(h,"") for h in auto_hdrs] for a in autos]
    _write_sheet(ws4, auto_hdrs, auto_rows)

    buf = __import__("io").BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ═══════════════════════════════════════════════════════════════════
#  PREMIUM TREND (helper — used from Analytics page)
# ═══════════════════════════════════════════════════════════════════
def _render_premium_trend(data):
    pols = data["policies"]
    from collections import defaultdict
    by_year = defaultdict(float)
    by_year_carrier = defaultdict(lambda: defaultdict(float))
    for p in pols:
        yr_str = str(p.get("effective_date",""))[:4]
        try:
            yr = int(yr_str)
            if 2000 <= yr <= 2035:
                prem = p.get("premium") or 0
                carrier = (p.get("carrier") or "Unknown")[:25]
                by_year[yr] += prem
                by_year_carrier[yr][carrier] += prem
        except Exception:
            pass

    if not by_year:
        st.info("No effective_date data available for trend chart.")
        return

    years = sorted(by_year.keys())
    carriers_all = sorted({c for yr in years for c in by_year_carrier[yr]})
    colors = ["#D4547A","#4A1942","#F59E0B","#22C55E","#60A5FA","#A78BFA","#FB923C","#34D399"]

    traces = []
    for i, carrier in enumerate(carriers_all):
        vals = [by_year_carrier[yr].get(carrier, 0) for yr in years]
        traces.append(go.Bar(
            name=carrier, x=[str(y) for y in years], y=vals,
            marker_color=colors[i % len(colors)],
        ))

    fig = go.Figure(data=traces)
    fig.update_layout(
        barmode="stack",
        plot_bgcolor=R['bg'], paper_bgcolor=R['bg'],
        font=dict(color=R['text_d'], size=11),
        margin=dict(l=0,r=0,t=30,b=0),
        height=320,
        legend=dict(orientation="h", x=0, y=1.1, font_size=10),
        title=dict(text="Total Premium by Year", font_size=12, x=0),
        yaxis=dict(tickprefix="$", gridcolor=R['rose_lt']),
        xaxis=dict(gridcolor=R['rose_lt']),
    )
    st.plotly_chart(fig, use_container_width=True, config={"displayModeBar": False})

    # YoY change
    if len(years) >= 2:
        rows_yoy = []
        for i in range(1, len(years)):
            prev, curr = by_year[years[i-1]], by_year[years[i]]
            chg = (curr - prev) / prev * 100 if prev else 0
            rows_yoy.append((years[i], f"${curr:,.0f}", f"{chg:+.1f}%"))
        st.markdown(f'<div class="sec-hdr">Year-over-Year Change</div>', unsafe_allow_html=True)
        for yr, total, chg in rows_yoy:
            color = "#dc2626" if chg.startswith("+") else "#16a34a"
            st.markdown(f"""
            <div style="display:flex;justify-content:space-between;padding:.3rem .6rem;
                 font-size:.82rem;border-bottom:1px solid {R['rose_lt']};">
              <span><b>{yr}</b></span><span>{total}</span>
              <span style="color:{color};font-weight:700;">{chg}</span>
            </div>""", unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════════
def main():
    if "logged_in" not in st.session_state:
        st.session_state.logged_in = False
    if "page" not in st.session_state:
        st.session_state.page = "dashboard"

    if not st.session_state.logged_in:
        show_login()
        return

    # Session timeout — auto-logout after 8 hours
    import time as _time
    if _time.time() - st.session_state.get("login_time", 0) > 8 * 3600:
        for k in list(st.session_state.keys()):
            del st.session_state[k]
        st.rerun()

    data = load_data()
    show_sidebar(data)

    page = st.session_state.get("page", "dashboard")
    if   page == "dashboard":   page_dashboard(data)
    elif page == "search":      page_search(data)
    elif page == "map":         page_map(data)
    elif page == "tasks":       page_tasks(data, save_data)
    elif page == "queue":       page_queue(data)
    elif page == "policies":    page_policies(data)
    elif page == "properties":  page_properties(data)
    elif page == "auto":        page_auto(data)
    elif page == "calendar":    page_calendar(data)
    elif page == "analytics":   page_analytics(data)
    elif page == "add":         page_add(data, save_data)
    elif page == "import_data": page_import(data, save_data)
    elif page == "reports":     page_reports(data)
    elif page == "settings":    page_settings(data, save_data)
    else:                       page_dashboard(data)


if __name__ == "__main__":
    main()
